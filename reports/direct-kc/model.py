# -*- coding: utf-8 -*-
"""Модель данных отчёта «Директ КЦ» за неделю 24.08–30.08.2026."""
import os
import openpyxl, re, json
from pathlib import Path
from collections import defaultdict, OrderedDict

U = Path(os.environ.get("SRC_DIR", "src"))   # каталог с исходными выгрузками
ADS  = U/"06a64ef5-project_294017_report_23_20260824_20260830.xlsx"
ADS2 = U/"15bb5749-project_294017_report_default_20260824_20260830.xlsx"
DEALS= U/"9917147d-list_orders_294017_20260823T21_00_0000_0020260830T20_59_5900_00.xlsx"

WEEK_START = "2026-08-24"
WEEK_END_EXCL = "2026-08-31"          # до 00:00 дня, следующего за концом недели
WEEK_LABEL = "24.08-30.08"

def norm(s):
    if s is None: return None
    s = str(s).replace('\xa0', ' ').strip()
    return s if s else None

def num(v):
    if v in (None, ''): return 0.0
    try: return float(str(v).replace('\xa0','').replace(' ','').replace(',','.'))
    except ValueError: return 0.0

# ------------------------------------------------------------------ реклама
def load_ads():
    wb = openpyxl.load_workbook(ADS, read_only=True, data_only=True)
    rows = list(wb["Report"].iter_rows(values_only=True)); wb.close()
    hdr = [norm(h) for h in rows[0]]; ix = {h: i for i, h in enumerate(hdr)}
    acc = defaultdict(lambda: defaultdict(float))
    camp = defaultdict(lambda: defaultdict(float))
    channels = defaultdict(set)
    stray = defaultdict(lambda: defaultdict(float))
    FIELDS = {'rashod':'Расходы','pokazy':'Показы','vizity':'Визиты',
              'zayavki':'Заявки','kval':'Квал-лид (пользовательский)',
              'lidkc':'Лид КЦ (пользовательский)',
              'vp_sozd':'ВП (реальная, с 1С) по дате создания (пользовательский)',
              'vp_prod':'ВП (реальная, с 1С) по дате продажи (пользовательский)',
              'prodazhi':'Продажи'}
    for r in rows[1:]:
        l1 = norm(r[0])
        if l1 is None or l1 == 'Итого/Среднее': continue
        l2, l3, l3id = norm(r[2]), norm(r[4]), norm(r[5])
        if not l1.startswith('ЯД '):
            for k, nm in FIELDS.items(): stray[l1][k] += num(r[ix[nm]])
            continue
        site = l1[3:].strip()
        cname = campaign_label(l2, l3)
        for k, nm in FIELDS.items():
            acc[site][k] += num(r[ix[nm]])
            camp[(site, cname)][k] += num(r[ix[nm]])
        if cname != NO_CAMP and l2 in ('Поиск', 'РСЯ'): channels[(site, cname)].add(l2)
        if cname != NO_CAMP and l3id: camp[(site, cname)]['_id'] = l3id
    return acc, camp, channels, stray

NO_CAMP = 'Не привязано к кампании'
_BAD_CAMP = {None, '—', 'type}', 'search', 'Нет значения'}
_BAD_CHAN = {'{source', 'Нет значения', 'Системные визиты'}
def campaign_label(l2, l3):
    if l3 in _BAD_CAMP or (l2 in _BAD_CHAN): return NO_CAMP
    return l3

# ------------------------------------------------------------------ сделки
SPAM_PAT = re.compile(r'спам|массов\w*\s*обзвон|рассылк|рекламн\w*\s*предложен|дубль|такой\s+лид\s*/\s*сделка\s+уже\s+есть', re.I)

def load_deals():
    wb = openpyxl.load_workbook(DEALS, read_only=True, data_only=True)
    rows = list(wb["Report"].iter_rows(values_only=True)); wb.close()
    hdr = [norm(h) for h in rows[0]]
    recs = []
    for r in rows[1:]:
        g = lambda i: norm(r[i-1])
        chain = g(9) or ''
        parts = [p.strip() for p in chain.split('→')]
        acc_site = parts[0][3:].strip() if parts and parts[0].startswith('ЯД ') else None
        chan = parts[1] if len(parts) > 1 else None
        camp = parts[2] if len(parts) > 2 else None
        cname = campaign_label(chan if chan in ('Поиск','РСЯ') else (chan or None), camp)
        if chan not in ('Поиск','РСЯ'): chan_out = chan or '—'
        else: chan_out = chan
        reasons = ' | '.join(x for x in (g(486), g(433), g(443), g(261)) if x)
        rec = dict(
            id=g(1), name=g(2), status=g(3), stage=g(27), date=g(5),
            src_check=g(35) or '', src_ctrl=g(459), chain=chain,
            account=acc_site, channel=chan_out, campaign=cname,
            summa=num(r[29]), vyruchka=num(r[3]),
            quality=(g(721) == 'Да'), reason=reasons,
            reason_lead=g(486), reason_deal=g(433),
            power_raw=g(551), power_group=g(701), goods=g(454),
            comment=g(33), responsible=g(34), roistat=g(189), utm_campaign=g(39),
        )
        rec['in_week'] = (WEEK_START <= (rec['date'] or '') < WEEK_END_EXCL + ' 00:00:00')
        rec['site_src'] = rec['src_check'].startswith('Сайт')
        rec['is_spam_dup'] = bool(SPAM_PAT.search(reasons)) if reasons else False
        rec['category'] = categorize(rec)
        rec['power'] = power_of(rec)
        rec['equipment'] = equipment_of(rec)
        recs.append(rec)
    return recs

LOSE_PAT = re.compile(r'\bLOSE\b|провален', re.I)
WIN_PAT  = re.compile(r'доставить\s+заказ|оплачено|успешн|\bWON\b', re.I)
def categorize(rec):
    if not rec['quality']: return None
    st = f"{rec['stage'] or ''} {rec['status'] or ''}"
    if LOSE_PAT.search(st): return 'провалена'
    if WIN_PAT.search(st):  return 'успешная'
    return 'в работе'

POWER_IN_NAME = re.compile(r'(\d{1,4}(?:[.,]\d)?)\s*(?:квт|kw|кв\b)', re.I)
def power_of(rec):
    p = rec['power_raw']
    if p:
        p = str(p).replace('.', ',').strip()
        return p
    m = POWER_IN_NAME.search(rec['name'] or '')
    if m: return m.group(1).replace('.', ',')
    return None

EQUIP_RULES = [
    (re.compile(r'фотосепаратор|сортировщик', re.I),                    'фотосепаратор'),
    (re.compile(r'\bМКС\b|модульн\w*\s*(?:компрессорн\w*\s*)?(?:станц|установ)|контейнерн\w*\s*станц', re.I), 'МКС'),
    (re.compile(r'азотн\w*\s*станц|генератор\w*\s*азота|\bазот\w*\b', re.I), 'генератор азота'),
    (re.compile(r'осушител', re.I),                                     'осушитель'),
    (re.compile(r'воздуходувк', re.I),                                  'воздуходувка'),
    (re.compile(r'дизель\w*|дизел\w+', re.I),                           'дизельный компрессор'),
    (re.compile(r'компрессор|винтов|безмасл|винт\.|сепарик|кубовик|ресивер', re.I), 'винтовой компрессор'),
]
def equipment_of(rec):
    for text in (rec['goods'], rec['name'], rec['comment']):
        if not text: continue
        for pat, label in EQUIP_RULES:
            if pat.search(text): return label
    return 'тип оборудования не указан'

# ------------------------------------------------------------------ сборка
def build():
    acc_ads, camp_ads, channels, stray = load_ads()
    recs = load_deals()

    included = [r for r in recs if r['in_week'] and r['site_src'] and r['account']]
    excluded = [r for r in recs if r not in included]

    def lead_flags(r):
        l1 = True
        l3 = r['quality']
        # запись, ставшая качественной сделкой, не считается спамом/дублем
        l2 = (not r['is_spam_dup']) or l3
        return l1, l2, l3

    acc_leads = defaultdict(lambda: [0, 0, 0])
    camp_leads = defaultdict(lambda: [0, 0, 0])
    for r in included:
        l1, l2, l3 = lead_flags(r)
        for i, f in enumerate((l1, l2, l3)):
            if f:
                acc_leads[r['account']][i] += 1
                camp_leads[(r['account'], r['campaign'])][i] += 1

    sites = set(acc_ads) | {r['account'] for r in included}
    def act(s):
        a = acc_ads.get(s, {})
        return (a.get('rashod', 0) or a.get('pokazy', 0) or a.get('vizity', 0)
                or acc_leads[s][0])
    sites = [s for s in sites if act(s)]
    sites.sort(key=lambda s: (-acc_ads.get(s, {}).get('rashod', 0), s))

    model = OrderedDict()
    for s in sites:
        a = acc_ads.get(s, defaultdict(float))
        L = acc_leads[s]
        deals = [r for r in included if r['account'] == s and r['quality']]
        cnames = {c for (ss, c) in camp_ads if ss == s} | {c for (ss, c) in camp_leads if ss == s}
        camps = []
        for c in cnames:
            ca = camp_ads.get((s, c), defaultdict(float)); CL = camp_leads[(s, c)]
            if not (ca.get('rashod', 0) or ca.get('pokazy', 0) or ca.get('vizity', 0) or CL[0]):
                continue
            camps.append(dict(name=c, ads=dict(ca), leads=CL,
                              channel='/'.join(sorted(channels.get((s, c), []))) or '—'))
        camps.sort(key=lambda c: (c['name'] == NO_CAMP, -c['ads'].get('rashod', 0), c['name']))
        model[s] = dict(ads=dict(a), leads=L, deals=deals, camps=camps)
    return model, recs, included, excluded, stray, camp_ads, camp_leads

if __name__ == '__main__':
    model, recs, included, excluded, stray, ca, cl = build()
    print("записей всего:", len(recs), "| включено:", len(included), "| исключено:", len(excluded))
    print("вне недели:", sum(1 for r in recs if not r['in_week']),
          "| не-сайтовый источник:", sum(1 for r in recs if r['in_week'] and not r['site_src']),
          "| без аккаунта ЯД:", sum(1 for r in recs if r['in_week'] and r['site_src'] and not r['account']))
    tot = [0, 0, 0]; trash = defaultdict(float)
    print(f"\n{'аккаунт':<28}{'расход':>10}{'показы':>10}{'визиты':>8}{'Л1':>4}{'Л2':>4}{'Л3':>4}{'ВПсозд':>8}{'ВПпрод':>9}{'прод':>5}  камп")
    for s, d in model.items():
        a = d['ads']; L = d['leads']
        for i in range(3): tot[i] += L[i]
        for k in ('rashod','pokazy','vizity','vp_sozd','vp_prod','prodazhi'): trash[k]+=a.get(k,0)
        print(f"{s:<28}{a.get('rashod',0):>10,.0f}{a.get('pokazy',0):>10,.0f}{a.get('vizity',0):>8,.0f}"
              f"{L[0]:>4}{L[1]:>4}{L[2]:>4}{a.get('vp_sozd',0):>8,.0f}{a.get('vp_prod',0):>9,.0f}{a.get('prodazhi',0):>5,.0f}  {len(d['camps'])}")
    print(f"{'ИТОГО':<28}{trash['rashod']:>10,.0f}{trash['pokazy']:>10,.0f}{trash['vizity']:>8,.0f}{tot[0]:>4}{tot[1]:>4}{tot[2]:>4}{trash['vp_sozd']:>8,.0f}{trash['vp_prod']:>9,.0f}{trash['prodazhi']:>5,.0f}")
    print("\nстороннее (не аккаунты ЯД):", {k: round(v['vizity']) for k, v in stray.items()})
    print("\nкатегории сделок:", dict((c, sum(1 for r in included if r['category'] == c)) for c in ('в работе','провалена','успешная')))
    print("исключённые записи:")
    for r in excluded:
        print("   ", r['id'], r['date'], "|", r['src_check'], "| качеств:", r['quality'], "| сумма:", r['summa'])
