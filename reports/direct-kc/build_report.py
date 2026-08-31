# -*- coding: utf-8 -*-
"""Сборка отчёта «Директ КЦ» за неделю 24.08–30.08.2026 в формате шаблона Metriki-OM-KTs_2026."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
import math
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from model import build, NO_CAMP, WEEK_LABEL
from textgen import summary_text

OUT = os.environ.get('OUT_XLSX', 'Отчет_Директ_КЦ_24.08-30.08.xlsx')

# ---- палитра и форматы шаблона -------------------------------------------
F_HEAD   = PatternFill('solid', fgColor='FFFFF2CC')   # шапка блока (св.-жёлтый)
F_WHITE  = PatternFill('solid', fgColor='FFFFFFFF')   # строки-источники
F_CALC   = PatternFill('solid', fgColor='FFE9D2E4')   # расчётные строки
F_NOTE   = PatternFill('solid', fgColor='FFFFF2CC')   # текстовый вывод
F_TABLE  = PatternFill('solid', fgColor='FFC9DAF8')   # шапки доп. листов
ARIAL    = 'Arial'
FMT_MONEY = '#,##0 ₽;[Red](#,##0 ₽);-'
FMT_MONEY2 = '#,##0.00 ₽;[Red](#,##0.00 ₽);-'
FMT_PCT   = '0.00%'
FMT_CNT   = '#,##0;[Red](#,##0);-'
THIN = Side(style='thin', color='FFD0D0D0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# порядок строк блока: (подпись, вид, формат)
ROWS = [
    ('Показы',              'src',  FMT_CNT),
    ('Визиты',              'src',  FMT_CNT),
    ('CTR',                 'calc', FMT_PCT),
    ('CPC',                 'calc', FMT_MONEY2),
    ('Конверсия в заявку',  'calc', FMT_PCT),
    ('Лид 1',               'src',  FMT_CNT),
    ('Стоимость Лида 1',    'calc', FMT_MONEY),
    ('Лиды 2',              'src',  FMT_CNT),
    ('Стоимость Лида 2',    'calc', FMT_MONEY),
    ('Рекламный бюджет',    'src',  FMT_MONEY),
    ('Лиды 3',              'src',  FMT_CNT),
    ('Стоимость Лида 3',    'calc', FMT_MONEY),
    ('ВП',                  'src',  FMT_MONEY),
    ('Продажи',             'src',  FMT_CNT),
    ('Средний чек',         'calc', FMT_MONEY),
]
I_POK, I_VIZ, I_CTR, I_CPC, I_CONV, I_L1, I_CL1, I_L2, I_CL2, I_BUD, I_L3, I_CL3, I_VP, I_PROD, I_CHECK = range(15)


def block(ws, top, values=None, sum_rows=None):
    """Пишет 15 стандартных строк блока начиная со строки `top`.
    values  — dict {индекс строки: число} для строк-источников;
    sum_rows — список номеров строк-первых-строк дочерних блоков (тогда источники = сумма)."""
    for i, (label, kind, fmt) in enumerate(ROWS):
        r = top + i
        a = ws.cell(r, 1, label)
        a.font = Font(name=ARIAL, sz=10, color='FF000000')
        a.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        a.fill = F_CALC if kind == 'calc' else F_WHITE
        a.border = BORDER
        c = ws.cell(r, 2)
        c.font = Font(name=ARIAL, sz=10)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.number_format = fmt
        c.fill = F_CALC if kind == 'calc' else F_WHITE
        c.border = BORDER
        if kind == 'calc':
            f = {
                I_CTR:   f'=IFERROR(B{top+I_VIZ}/B{top+I_POK},0)',
                I_CPC:   f'=IFERROR(B{top+I_BUD}/B{top+I_VIZ},0)',
                I_CONV:  f'=IFERROR(B{top+I_L1}/B{top+I_VIZ},0)',
                I_CL1:   f'=IFERROR(B{top+I_BUD}/B{top+I_L1},0)',
                I_CL2:   f'=IFERROR(B{top+I_BUD}/B{top+I_L2},0)',
                I_CL3:   f'=IFERROR(B{top+I_BUD}/B{top+I_L3},0)',
                I_CHECK: f'=IFERROR(B{top+I_VP}/B{top+I_PROD},0)',
            }[i]
            c.value = f
        elif sum_rows:
            c.value = '=' + '+'.join(f'B{s + i}' for s in sum_rows)
        else:
            c.value = (values or {}).get(i, 0)
    return top + 15


def header_row(ws, r, text, kind):
    """kind: 'total' | 'account' | 'sub' | 'campaign'"""
    a = ws.cell(r, 1, text)
    b = ws.cell(r, 2)
    if kind == 'total':
        for c in (a, b):
            c.fill = F_HEAD; c.font = Font(name=ARIAL, sz=10, b=True); c.border = BORDER
        b.value = f'=B$1'
    elif kind == 'account':
        a.font = Font(name=ARIAL, sz=11, b=True, color='FF0000FF')
    elif kind == 'sub':
        a.font = Font(name=ARIAL, sz=10, b=True)
        b.value = '=B$1'; b.font = Font(name=ARIAL, sz=10, b=True)
    else:
        a.font = Font(name=ARIAL, sz=10, b=True)
        b.value = '=B$1'; b.font = Font(name=ARIAL, sz=10, b=True)
    a.alignment = Alignment(horizontal='left', vertical='center')
    b.alignment = Alignment(horizontal='center', vertical='center')
    return r + 1


def note_block(ws, r, text):
    h = ws.cell(r, 1, 'Вывод по сделкам за неделю')
    h.font = Font(name=ARIAL, sz=10, b=True)
    h.fill = F_NOTE; h.alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(r, 2).fill = F_NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    t = ws.cell(r, 1, text)
    t.font = Font(name=ARIAL, sz=10)
    t.fill = F_NOTE
    t.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.cell(r, 2).fill = F_NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = max(30, 15.0 * math.ceil(len(text) / 110))
    return r + 1


def main():
    model, recs, included, excluded, stray, camp_ads, camp_leads = build()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Директ КЦ'
    ws.sheet_properties.tabColor = 'FFFFF2CC'

    # ---- шапка: дата недели в верхней горизонтальной шапке столбца B ------
    h = ws.cell(1, 2, WEEK_LABEL)
    h.font = Font(name=ARIAL, sz=10, b=True)
    h.alignment = Alignment(horizontal='center', vertical='center')
    h.border = BORDER
    ws.column_dimensions['A'].width = 55.3
    ws.column_dimensions['B'].width = 18
    ws.freeze_panes = 'B2'
    ws.print_title_rows = '1:1'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ---- общий блок (значения проставим формулами позже) -----------------
    r = 2
    total_hdr = r
    r = header_row(ws, r, 'Яндекс Директ Общие', 'total')
    total_top = r
    r += 15                                   # заполним после аккаунтов
    r += 1
    all_deals = [x for x in included if x['quality']]
    r = note_block(ws, r, summary_text(all_deals))
    r += 1

    account_tops = []
    for site, d in model.items():
        r = header_row(ws, r, site.upper(), 'account')
        r = header_row(ws, r, 'Общие', 'sub')
        acc_top = r
        account_tops.append(acc_top)
        camp_tops = []
        # сначала считаем, где будут блоки кампаний (аккаунт = сумма кампаний)
        r += 15
        r += 1
        r = note_block(ws, r, summary_text(d['deals']))
        r += 1
        for c in d['camps']:
            label = c['name']
            # префикс канала добавляем только если в названии РК его ещё нет
            if label != NO_CAMP and c['channel'] != '—' and not label.startswith('['):
                label = f"[{c['channel'].replace('/', ' и ')}] {label}"
            r = header_row(ws, r, label, 'campaign')
            ct = r
            camp_tops.append(ct)
            a = c['ads']; L = c['leads']
            block(ws, ct, values={
                I_POK: round(a.get('pokazy', 0)), I_VIZ: round(a.get('vizity', 0)),
                I_L1: L[0], I_L2: L[1], I_L3: L[2],
                I_BUD: a.get('rashod', 0),
                I_VP: a.get('vp_sozd', 0),
                I_PROD: sum(1 for x in d['deals']
                            if x['campaign'] == c['name'] and x['category'] == 'успешная'),
            })
            r = ct + 15 + 1
        block(ws, acc_top, sum_rows=camp_tops)
        r += 1
    block(ws, total_top, sum_rows=account_tops)

    # ================================================== лист «Сделки»
    ds = wb.create_sheet('Сделки')
    cols = ['ID сделки', 'Название', 'Сайт', 'РК', 'Поиск/РСЯ', 'Дата создания',
            'Текущий статус', 'Категория', 'Сумма сделки', 'Мощность',
            'Группа мощности', 'Тип оборудования', 'Источник проверки',
            'Причина провала', 'Ответственный']
    widths = [14, 46, 22, 34, 11, 18, 42, 12, 14, 10, 24, 22, 26, 40, 26]
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ds.cell(1, i, c)
        cell.font = Font(name=ARIAL, sz=10, b=True)
        cell.fill = F_TABLE; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ds.column_dimensions[get_column_letter(i)].width = w
    deals = sorted((x for x in included if x['quality']), key=lambda x: x['date'])
    rr = 2
    for x in deals:
        vals = [x['id'], x['name'], x['account'],
                x['campaign'], x['channel'], x['date'], x['status'], x['category'],
                x['summa'], x['power'] or '—', x['power_group'] or '—',
                x['equipment'], x['src_check'],
                x['reason'] or '—', x['responsible']]
        for i, v in enumerate(vals, 1):
            cell = ds.cell(rr, i, v)
            cell.font = Font(name=ARIAL, sz=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=(i in (2, 7, 14)))
            if i == 9: cell.number_format = FMT_MONEY
        rr += 1
    t = ds.cell(rr, 1, f'ИТОГО сделок: {len(deals)}')
    t.font = Font(name=ARIAL, sz=10, b=True)
    ds.cell(rr, 8, '').font = Font(name=ARIAL, sz=10, b=True)
    s = ds.cell(rr, 9, f'=SUM(I2:I{rr-1})')
    s.font = Font(name=ARIAL, sz=10, b=True); s.number_format = FMT_MONEY
    for i in range(1, 16): ds.cell(rr, i).fill = F_TABLE; ds.cell(rr, i).border = BORDER
    rr += 2
    for line in [
        'Категории: «в работе» — активные стадии; «провалена» — стадия LOSE / статус «Провалена»;',
        '«успешная» — «Доставить заказ», «Оплачено», «Успешно».',
        'Сумма сделки — предварительная оценка воронки, это НЕ ВП и НЕ подтверждённая выручка.',
    ]:
        ds.cell(rr, 1, line).font = Font(name=ARIAL, sz=9, i=True)
        rr += 1
    ds.freeze_panes = 'A2'
    ds.page_setup.orientation = 'landscape'
    ds.page_setup.fitToWidth = 1
    ds.page_setup.fitToHeight = 0
    ds.sheet_properties.pageSetUpPr.fitToPage = True
    ds.auto_filter.ref = f'A1:O{rr-4}'

    # ================================================== лист «Комментарий»
    cs = wb.create_sheet('Комментарий')
    cs.page_setup.orientation = 'landscape'
    cs.page_setup.fitToWidth = 1
    cs.page_setup.fitToHeight = 0
    cs.sheet_properties.pageSetUpPr.fitToPage = True
    cs.column_dimensions['A'].width = 16
    for col, w in zip('BCDEFGHI', [18, 26, 34, 24, 30, 40, 14, 46]):
        cs.column_dimensions[col].width = w
    by_src = [x for x in recs if x['in_week'] and not x['site_src']]
    by_date = [x for x in recs if not x['in_week']]
    no_camp = [x for x in included if x['quality'] and x['campaign'] == NO_CAMP]
    excl_sum = sum(x['summa'] for x in by_src + by_date)
    rr = 1
    title = cs.cell(rr, 1, f'Исключения и оговорки — отчётная неделя {WEEK_LABEL}.2026')
    title.font = Font(name=ARIAL, sz=12, b=True); rr += 2
    for label, val in [
        ('Всего записей в выгрузке сделок', len(recs)),
        ('Включено в отчёт (Лид 1)', len(included)),
        ('Исключено записей, всего', len(by_src) + len(by_date)),
        ('— по не-сайтовому источнику проверки', len(by_src)),
        ('— по дате вне отчётной недели', len(by_date)),
        ('Сумма исключённых сделок, ₽', excl_sum),
        ('Сделок Лид 3 без определённой РК', len(no_camp)),
    ]:
        a = cs.cell(rr, 1, label); a.font = Font(name=ARIAL, sz=10, b=True)
        cs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
        b = cs.cell(rr, 3, val); b.font = Font(name=ARIAL, sz=10, b=True)
        b.number_format = FMT_MONEY if 'Сумма' in label else FMT_CNT
        b.alignment = Alignment(horizontal='left')
        rr += 1
    rr += 1

    hdr = ['ID', 'Дата создания', 'Источник проверки', 'Исходный источник',
           'Аккаунт/сайт', 'Кампания', 'Статус', 'Сумма', 'Причина исключения']

    def table_header(rr):
        for i, c in enumerate(hdr, 1):
            cell = cs.cell(rr, i, c)
            cell.font = Font(name=ARIAL, sz=10, b=True); cell.fill = F_TABLE
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return rr + 1

    def table_rows(rr, items, reason):
        if not items:
            cs.cell(rr, 1, 'Записей нет').font = Font(name=ARIAL, sz=10, i=True)
            return rr + 1
        for x in items:
            vals = [x['id'], x['date'], x['src_check'] or '—', x['chain'] or '—',
                    x['account'] or '—', x['campaign'] or '—', x['status'],
                    x['summa'], reason(x)]
            for i, v in enumerate(vals, 1):
                cell = cs.cell(rr, i, v)
                cell.font = Font(name=ARIAL, sz=10); cell.border = BORDER
                cell.alignment = Alignment(vertical='top', wrap_text=(i in (4, 7, 9)))
                if i == 8: cell.number_format = FMT_MONEY
            rr += 1
        return rr

    s1 = cs.cell(rr, 1, '1. Исключены: источник проверки не начинается со слова «Сайт»')
    s1.font = Font(name=ARIAL, sz=11, b=True); rr += 1
    rr = table_header(rr)
    rr = table_rows(rr, by_src, lambda x: f'Источник проверки «{x["src_check"]}» — не сайтовый трафик, вне отчёта по Директу')
    rr += 1
    s2 = cs.cell(rr, 1, '2. Исключены: дата создания вне отчётной недели')
    s2.font = Font(name=ARIAL, sz=11, b=True); rr += 1
    rr = table_header(rr)
    rr = table_rows(rr, by_date, lambda x: 'Дата создания вне периода 24.08.2026 00:00 — 31.08.2026 00:00')
    rr += 1
    s3 = cs.cell(rr, 1, '3. Учтены в отчёте, но рекламная кампания не определена')
    s3.font = Font(name=ARIAL, sz=11, b=True); rr += 1
    rr = table_header(rr)
    rr = table_rows(rr, no_camp, lambda x: 'В цепочке «Источник» нет РК; roistat/utm_campaign не содержат кампании — блок «Не привязано к кампании»')
    rr += 2

    s4 = cs.cell(rr, 1, 'Справочно (записи НЕ исключены из отчёта)')
    s4.font = Font(name=ARIAL, sz=11, b=True); rr += 1
    mism = []
    for x in included:
        site = x['src_check'].replace('Сайт', '').strip()
        if x['account'] and site and site != x['account'] and not site.endswith(x['account']):
            mism.append(x)
    notes = [
        f'Расхождение сайта в источнике проверки и в цепочке «Источник»: {len(mism)} записей '
        f'({", ".join(x["id"] for x in mism)}). Отбор в отчёт выполнен по источнику проверки, '
        f'привязка к аккаунту и РК — по цепочке «Источник» (приоритет 1 правила привязки).',
        f'Визиты Яндекс Директа без привязки к аккаунту (сломанная UTM-разметка: '
        f'{", ".join(sorted(stray))}): {round(sum(v["vizity"] for v in stray.values()))} визитов, '
        f'0 ₽ расхода, 0 заявок. В блоки аккаунтов не включены.',
        'ВП по всем аккаунтам = 0: поля «ВП (маржа)», «ВП (реальная, с 1С)», «ВП (ориентировочная)» '
        'в выгрузке сделок не заполнены, а «ВП (реальная, с 1С) по дате создания» в выгрузке Roistat '
        'равна нулю за неделю. ROI не рассчитывается.',
    ]
    for n in notes:
        c = cs.cell(rr, 1, n)
        c.font = Font(name=ARIAL, sz=10)
        c.alignment = Alignment(vertical='top', wrap_text=True)
        cs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
        cs.row_dimensions[rr].height = 45
        rr += 1

    # ================================================== лист «Методика»
    ms = wb.create_sheet('Методика')
    ms.page_setup.orientation = 'landscape'
    ms.page_setup.fitToWidth = 1
    ms.page_setup.fitToHeight = 0
    ms.sheet_properties.pageSetUpPr.fitToPage = True
    ms.column_dimensions['A'].width = 34
    ms.column_dimensions['B'].width = 122
    rr = 1
    t = ms.cell(rr, 1, f'Методика расчёта — отчёт «Директ КЦ», неделя {WEEK_LABEL}.2026')
    t.font = Font(name=ARIAL, sz=12, b=True); rr += 2
    METHOD = [
        ('Отчётный период',
         'С 24.08.2026 00:00 до 31.08.2026 00:00 (включительно 24.08–30.08.2026). Период определён по именам выгрузок '
         'Roistat (20260824_20260830) и по диапазону выгрузки сделок. Дата включения — поле «Дата» (дата создания '
         'лида/сделки). Записи вне периода в показатели не входят и вынесены на лист «Комментарий».'),
        ('Источник проверки (главный фильтр)',
         'Используется второе поле «Источник» выгрузки сделок («Источник.1», в Битрикс24 — проверенный менеджером источник; '
         'дублируется полем «*источник - контроль»). В отчёт попадают только записи, значение которых начинается со слова '
         '«Сайт» (Сайт prokompressor.ru, Сайт enger-air.ru и т. д.). «Постоянный с интернета», «Звонок (не используем)», '
         'SEO, прямой заход, e-mail, рассылки, мессенджеры, офлайн и неопределённый источник исключаются, даже если в поле '
         '«Источник» указано «ЯД …» или в UTM есть yandex.'),
        ('Лид 1',
         'Все обращения Яндекс Директа, созданные в отчётную неделю, прошедшие фильтр источника проверки «Сайт …» и '
         'привязанные к аккаунту Яндекс Директа (уровень 1 цепочки «Источник» вида «ЯД сайт»). За неделю — 145 записей.'),
        ('Лид 2',
         'Лид 1 за вычетом записей, у которых в причинах закрытия («Причина провала лида (КЦ)», «Причина провала сделки '
         '(КЦ-Продажа)» и текстовых полях причин) встречается: «Спам», «массовый обзвон», «рассылка», «рекламные '
         'предложения», «Дубль», «Такой лид/сделка уже есть». Причины «не наш ассортимент», «нужна консультация», '
         '«нет потребности», «не выходит на связь», «торговая организация», «купили у конкурента» НЕ исключаются. '
         'За неделю — 98 записей.'),
        ('Оговорка по Лид 2',
         'Две записи (deal_154334, deal_154036) имеют причину «Дубль. Такой лид/сделка уже есть», но при этом являются '
         'качественными лидами с созданной сделкой. Они оставлены в Лид 2, чтобы воронка оставалась монотонной '
         '(Лид 1 ≥ Лид 2 ≥ Лид 3). Без этой оговорки Лид 2 составил бы 96.'),
        ('Лид 3',
         'Качественный лид, равный созданной сделке: поле «Качественный лид» = «Да» и одновременно заполнены ID вида '
         'deal_…, стадия и статус сделки. Двойного учёта лида и созданной из него сделки нет — в выгрузке это одна '
         'запись. За неделю — 28 сделок.'),
        ('Привязка к РК',
         'Приоритет: 1) цепочка поля «Источник» («ЯД сайт → Поиск/РСЯ → Кампания»); 2) значение roistat; 3) utm_campaign; '
         '4) идентификатор РК; 5) рекламная выгрузка Roistat. Если сайт определён, а конкретная РК — нет, лид и сделка '
         'учитываются в итогах сайта и показываются в блоке «Не привязано к кампании»; кампания по догадке не '
         'присваивается. Одна кампания, работавшая и в Поиске, и в РСЯ, показана одним блоком с префиксом '
         '«[Поиск и РСЯ]».'),
        ('Статусы сделок',
         '«В работе» — незакрытые стадии (выявление потребности, подготовка и отправка КП, согласование КП, счёт, договор, '
         'предоплата, бронирование, логистика). «Провалена» — стадия LOSE или статус «Провалена». «Успешная» — '
         '«Доставить заказ», «Оплачено», «Успешно». В табличные блоки эти статусы не выводятся — только в текстовый вывод '
         'и на лист «Сделки».'),
        ('Мощность',
         'Приоритет: 1) «Мощность мах ком-ра в сделке»; 2) «Мощность оборудования»; 3) название сделки, если мощность в нём '
         'указана однозначно (7,5 кВт, 15 кВт, 37 кВт, 110 кВт). Одинаковые мощности сгруппированы. Поле «Группа по '
         'мощности» (до 37 кВт / от 45 кВт) вынесено на лист «Сделки» справочно.'),
        ('Оборудование',
         'Приоритет: 1) товарные позиции (поле «Группа товара»); 2) название сделки; 3) комментарий; 4) название РК — '
         'только как вспомогательный признак. Нормализация: МКС / модульная (компрессорная) станция / контейнерная '
         'станция → МКС; азотная станция / генератор азота / азот → генератор азота; осушитель любого типа → осушитель; '
         'дизель, дизельный → дизельный компрессор; воздуходувка → воздуходувка; компрессор, винтовой компрессор → '
         'винтовой компрессор (если нет более точного типа); фотосепаратор, сортировщик → фотосепаратор. Если тип '
         'определить нельзя — «тип оборудования не указан».'),
        ('Ориентировочная сумма активных сделок',
         'Сумма поля «Сумма сделки» только по сделкам в статусе «В работе». Это предварительная денежная оценка активной '
         'воронки: НЕ ВП, НЕ маржа и НЕ подтверждённая выручка. Слово «выручка» к активным сделкам не применяется.'),
        ('ВП и продажи',
         'ВП берётся только из полей валовой прибыли («ВП», «ВП (маржа)», «ВП (реальная, с 1С)»). За отчётную неделю они '
         'не заполнены, а «ВП (реальная, с 1С) по дате создания» в выгрузке Roistat равна 0 — поэтому строка «ВП» = 0, '
         'ROI не рассчитывается, средний чек = 0. Сумма сделки ВП не подменяет. В строку «Продажи» включена только '
         'успешно завершённая сделка (1 шт., ironmac-compressor.com).'),
        ('Источники данных',
         'Показы, визиты, расходы — выгрузка Roistat project_294017_report_23_20260824_20260830.xlsx (фильтр Яндекс '
         'Директа). Лид 1 / Лид 2 / Лид 3, статусы, суммы, мощности, оборудование — выгрузка сделок '
         'list_orders_294017_20260823T21:00–20260830T20:59. Вторая выгрузка Roistat (report_default) использована для '
         'сверки заявок и квал-лидов. Оформление — лист «Директ КЦ» файла Metriki-OM-KTs_2026.xlsx.'),
        ('Формулы',
         'CTR = Визиты / Показы; CPC = Рекламный бюджет / Визиты; Конверсия в заявку = Лид 1 / Визиты; '
         'Стоимость Лида 1 = Рекламный бюджет / Лид 1; Стоимость Лида 2 = Рекламный бюджет / Лиды 2; '
         'Стоимость Лида 3 = Рекламный бюджет / Лиды 3; Средний чек = ВП / Продажи. Все расчётные строки обёрнуты в '
         'IFERROR(…;0). Блок аккаунта — сумма блоков его кампаний, блок «Яндекс Директ Общие» — сумма блоков аккаунтов.'),
    ]
    for k, v in METHOD:
        a = ms.cell(rr, 1, k)
        a.font = Font(name=ARIAL, sz=10, b=True)
        a.fill = F_TABLE; a.border = BORDER
        a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        b = ms.cell(rr, 2, v)
        b.font = Font(name=ARIAL, sz=10); b.border = BORDER
        b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ms.row_dimensions[rr].height = max(30, 13.5 * (len(v) // 118 + 1))
        rr += 1

    wb.calculation.fullCalcOnLoad = True
    wb.save(OUT)
    print('OK ->', OUT)
    print('аккаунтов:', len(model), '| блоков кампаний:', sum(len(d['camps']) for d in model.values()))
    print('строк на листе «Директ КЦ»:', ws.max_row)


if __name__ == '__main__':
    main()
