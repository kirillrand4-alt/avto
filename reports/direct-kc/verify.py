# -*- coding: utf-8 -*-
import os, re, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
import openpyxl
from model import build, NO_CAMP

CALC = os.environ.get('CALC_XLSX', 'recalc/Отчет_Директ_КЦ_24.08-30.08.xlsx')
SRC  = os.environ.get('SRC_XLSX', 'Отчет_Директ_КЦ_24.08-30.08.xlsx')
ERRS = ('#DIV/0!', '#REF!', '#VALUE!', '#NAME?', '#N/A', '#NULL!', '#NUM!', 'Err:')
ok = True
def chk(cond, msg):
    global ok
    print(('  OK  ' if cond else '  FAIL') + '  ' + msg)
    if not cond: ok = False

wb = openpyxl.load_workbook(CALC, data_only=True)
wbf = openpyxl.load_workbook(SRC)
model, recs, included, excluded, stray, camp_ads, camp_leads = build()

print("== 1. структура книги ==")
chk(wb.sheetnames[0] == 'Директ КЦ', f"первый лист = «Директ КЦ» (получено {wb.sheetnames})")
chk(set(wb.sheetnames) == {'Директ КЦ','Сделки','Комментарий','Методика'}, "состав листов")

print("== 2. ошибок формул нет ни на одном листе ==")
bad = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and any(e in c.value for e in ERRS):
                bad.append(f"{ws.title}!{c.coordinate}={c.value}")
chk(not bad, f"формульных ошибок: {len(bad)} {bad[:5]}")

print("== 3. пустых (невычисленных) формул нет ==")
ws, wsf = wb['Директ КЦ'], wbf['Директ КЦ']
empt = [wsf.cell(r,2).coordinate for r in range(1, wsf.max_row+1)
        if isinstance(wsf.cell(r,2).value, str) and wsf.cell(r,2).value.startswith('=')
        and ws.cell(r,2).value is None]
chk(not empt, f"невычисленных формул: {len(empt)} {empt[:5]}")

print("== 4. шапка недели ==")
chk(ws['B1'].value == '24.08-30.08', f"B1 = {ws['B1'].value!r}")
chk(ws['A1'].value is None, "A1 пуст — период только в верхней шапке столбца")
per = [f"A{r}" for r in range(1, ws.max_row+1) if str(ws.cell(r,1).value or '').strip().lower() == 'период']
chk(not per, f"отдельного столбца/строки «Период» нет {per}")

print("== 5. запрещённые строки в табличных блоках ==")
FORBID = ['сделок в работе','провалено сделок','успешно сделок','сумма всех сделок',
          'потенциал сделок','сумма проваленных сделок','выручка успешных сделок']
found = []
for r in range(1, ws.max_row+1):
    v = str(ws.cell(r,1).value or '').strip().lower()
    if v in FORBID: found.append(f"A{r}={v}")
chk(not found, f"запрещённых строк: {found}")

print("== 6. состав строк каждого блока ==")
EXPECT = ['Показы','Визиты','CTR','CPC','Конверсия в заявку','Лид 1','Стоимость Лида 1','Лиды 2',
          'Стоимость Лида 2','Рекламный бюджет','Лиды 3','Стоимость Лида 3','ВП','Продажи','Средний чек']
tops = [r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == 'Показы']
badb = [t for t in tops if [ws.cell(t+i,1).value for i in range(15)] != EXPECT]
chk(not badb, f"блоков всего {len(tops)}, некорректных {len(badb)}")

print("== 7. сверка «Яндекс Директ Общие» с исходными данными ==")
top = tops[0]
exp = dict(pok=sum(d['ads'].get('pokazy',0) for d in model.values()),
           viz=sum(d['ads'].get('vizity',0) for d in model.values()),
           bud=sum(d['ads'].get('rashod',0) for d in model.values()),
           l1=sum(d['leads'][0] for d in model.values()),
           l2=sum(d['leads'][1] for d in model.values()),
           l3=sum(d['leads'][2] for d in model.values()))
got = dict(pok=ws.cell(top,2).value, viz=ws.cell(top+1,2).value, l1=ws.cell(top+5,2).value,
           l2=ws.cell(top+7,2).value, bud=ws.cell(top+9,2).value, l3=ws.cell(top+10,2).value)
for k in exp:
    chk(abs((got[k] or 0) - exp[k]) < 1.0, f"{k}: в файле {got[k]!r}, ожидалось {round(exp[k],2)}")
chk(got['l1'] == 145 and got['l2'] == 98 and got['l3'] == 28, "Лид1=145, Лид2=98, Лид3=28")
chk(ws.cell(top+12,2).value in (0, None), f"ВП = {ws.cell(top+12,2).value} (не заполнена в источниках)")
chk(ws.cell(top+13,2).value == 1, f"Продажи = {ws.cell(top+13,2).value} (1 успешная сделка)")

print("== 8. расчётные строки пересчитаны верно ==")
def near(a,b,t=1e-6): return abs((a or 0)-(b or 0)) <= t*max(1,abs(b or 1))
c = ws.cell
prob = []
for t in tops:
    pok,viz,l1,l2,bud,l3,vp,prod = (c(t+i,2).value or 0 for i in (0,1,5,7,9,10,12,13))
    for i, expv in [(2, viz/pok if pok else 0), (3, bud/viz if viz else 0), (4, l1/viz if viz else 0),
                    (6, bud/l1 if l1 else 0), (8, bud/l2 if l2 else 0), (11, bud/l3 if l3 else 0),
                    (14, vp/prod if prod else 0)]:
        if not near(c(t+i,2).value, expv, 1e-9): prob.append((t+i, c(t+i,2).value, expv))
chk(not prob, f"расхождений в расчётных строках: {len(prob)} {prob[:3]}")

print("== 9. аккаунт = сумма его кампаний; общий = сумма аккаунтов ==")
accs = [r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == 'Общие']
chk(len(accs) == len(model), f"блоков «Общие» (аккаунтов): {len(accs)} из {len(model)}")
mism = []
for site, d in model.items():
    a = d['ads']; L = d['leads']
    hdr_r = next(r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == site.upper())
    t = hdr_r + 2
    for i, expv in [(0,a.get('pokazy',0)),(1,a.get('vizity',0)),(5,L[0]),(7,L[1]),(9,a.get('rashod',0)),(10,L[2])]:
        if abs((ws.cell(t+i,2).value or 0) - expv) > 0.005:
            mism.append((site, EXPECT[i], ws.cell(t+i,2).value, expv))
chk(not mism, f"расхождений аккаунт/Roistat: {len(mism)} {mism[:3]}")

print("== 10. текстовый вывод под каждым аккаунтом ==")
notes = [r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == 'Вывод по сделкам за неделю']
chk(len(notes) == len(model)+1, f"блоков вывода: {len(notes)} (аккаунтов {len(model)} + общий)")
noacc = []
for site, d in model.items():
    hdr_r = next(r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == site.upper())
    nxt = [n for n in notes if n > hdr_r]
    if not nxt: noacc.append(site); continue
    txt = ws.cell(nxt[0]+1, 1).value or ''
    n = len(d['deals'])
    m = re.search(r'создан[оа] (\d+)', txt)
    cnt = int(m.group(1)) if m else (0 if 'не создано' in txt else -1)
    if cnt != n: noacc.append((site, cnt, n))
    if 'выручк' in txt.lower(): noacc.append((site, 'слово «выручка» в тексте активных сделок'))
chk(not noacc, f"несоответствий текст/Лид3: {noacc}")

print("== 11. в отчёт не попали не-сайтовые источники и записи вне недели ==")
ds = wb['Сделки']
ids = {ds.cell(r,1).value for r in range(2, ds.max_row+1) if str(ds.cell(r,1).value or '').startswith('deal_')}
badsrc = {x['id'] for x in recs if not x['site_src']} & ids
baddate = {x['id'] for x in recs if not x['in_week']} & ids
chk(not badsrc, f"не-сайтовых записей на листе «Сделки»: {badsrc}")
chk(not baddate, f"записей вне недели: {baddate}")
chk(len(ids) == 28, f"сделок на листе «Сделки»: {len(ids)} (ожидалось 28)")

print("== 12. в работе + провалено + успешно = число сделок ==")
cats = [ds.cell(r,8).value for r in range(2, ds.max_row+1) if str(ds.cell(r,1).value or '').startswith('deal_')]
from collections import Counter
cc = Counter(cats)
chk(cc['в работе']+cc['провалена']+cc['успешная'] == 28, f"{dict(cc)} — сумма {sum(cc.values())}")

print("== 13. лист «Комментарий» ==")
cs = wb['Комментарий']
vals = {str(cs.cell(r,1).value or ''): cs.cell(r,3).value for r in range(1, 20)}
chk(any('Исключено записей, всего' in k for k in vals), "есть итог по исключениям")
txt_all = ' '.join(str(cs.cell(r,ccc).value or '') for r in range(1, cs.max_row+1) for ccc in range(1,10))
for i in ('lead_372004','deal_154616','deal_154588','deal_154465','deal_154303'):
    chk(i in txt_all, f"{i} присутствует на листе «Комментарий»")

print()
print("ИТОГ:", "ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ" if ok else "ЕСТЬ ЗАМЕЧАНИЯ")
