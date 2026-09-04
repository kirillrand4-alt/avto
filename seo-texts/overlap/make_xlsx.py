#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Книга Excel с разбором пересечения сайтов по запросам.

Числа пишутся числами: в CSV русский Excel превращал позицию 17.6 в дату
17 июня. Формул в книге нет намеренно — это выгрузка из внешнего набора
данных, а не модель; пересчитывать нечего, зато файл читается любым
инструментом без предварительного открытия в Excel.
"""
import sys, os, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_overlap as bo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "."
OUT = sys.argv[2] if len(sys.argv) > 2 else "peresechenie-saytov.xlsx"
N = 15
FONT, HEAD = Font(name="Arial", size=10), Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="404040")
NOTE = Font(name="Arial", size=10, italic=True, color="808080")


def sheet(wb, title, headers, rows, widths, formats=None, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill = HEAD, FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 30
    for r in rows:
        ws.append(r)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = FONT
    for col, fmt in (formats or {}).items():
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = fmt
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return ws


def main():
    rows = bo.load_q_files(SRC)
    agg = bo.aggregate(rows)
    agg_ya, agg_go = bo.aggregate(rows, "Яндекс"), bo.aggregate(rows, "Google")
    idx = bo.site_index(agg)
    sites = sorted(idx, key=lambda s: -sum(v[0] for v in idx[s].values()))
    owners = collections.defaultdict(set)
    for (d, q) in agg:
        owners[q].add(d)
    top = {s: [q for q, _ in sorted(idx[s].items(), key=lambda kv: -kv[1][0])[:N]] for s in sites}
    tset = {s: set(v) for s, v in top.items()}
    aset = {s: set(idx[s]) for s in sites}

    def wpos(site, qs, a):
        i = sum(a[(site, q)][0] for q in qs if (site, q) in a)
        return None if not i else sum(a[(site, q)][2] * a[(site, q)][0] for q in qs if (site, q) in a) / i

    wb = Workbook()
    wb.remove(wb.active)

    # --- лист «Источник» ---
    ws = wb.create_sheet("Источник")
    tot_i = sum(v[0] for v in agg.values())
    tot_c = sum(v[1] for v in agg.values())
    for line in [
        ("Пересечение сайтов по поисковым запросам", True),
        ("", False),
        ("Данные: выгрузка запросов с дропа, файлы q_<домен>.txt", False),
        ("Период: 2026-05-25 — 2026-08-24. Поисковики: Яндекс и Google.", False),
        (f"Доменов: {len(sites)}. Уникальных запросов: {len(owners):,}. "
         f"Показов: {tot_i:,}. Кликов: {tot_c:,}.", False),
        ("", False),
        ("Запросы нормализованы: регистр, лишние пробелы, ё → е.", False),
        ("Позиция — средняя за период, взвешенная по показам.", False),
        ("«Топ-15» — 15 запросов сайта с наибольшими показами (Яндекс + Google).", False),
        ("", False),
        ("Оговорки к данным:", True),
        ("1. Клики berg-compressor.com и berg-kompressor.ru за период завышены "
         "накруткой ПФ (выключена 13.07.2026). По недельному срезу берг: до 13.07 "
         "CTR 30%, после — 5,9%; ботом было ~95% кликов и ~72% показов. "
         "Клики этой пары для выводов не использовать.", False),
        ("2. Показатели Google ненадёжны: CTR в топ-3 составляет 0,93% против "
         "15,96% у Яндекса при 65% всех показов сети. Похоже на парсеры позиций. "
         "Решения принимать по яндексовому срезу.", False),
        ("", False),
        ("Все числа в книге — значения, посчитанные скриптами из "
         "seo-texts/overlap/ репозитория avto. Формул в книге нет: это выгрузка "
         "из внешнего набора данных, пересчитывать внутри книги нечего. "
         "Пересобрать: python3 make_xlsx.py <каталог с q_*.txt> <файл.xlsx>", False),
    ]:
        ws.append([line[0]])
        ws.cell(ws.max_row, 1).font = Font(name="Arial", size=11, bold=True) if line[1] else FONT
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120

    # --- Топ-15 позиции ---
    data = []
    for s in sites:
        qs = top[s]
        i = sum(idx[s][q][0] for q in qs)
        tot = sum(v[0] for v in idx[s].values())
        c = sum(idx[s][q][1] for q in qs)
        py, pg = wpos(s, qs, agg_ya), wpos(s, qs, agg_go)
        shared = sum(1 for q in qs if len(owners[q]) > 1)
        nb = len({o for o in sites if o != s and tset[s] & aset[o]})
        data.append([s, i, tot, i / tot if tot else 0, c, round(py, 1) if py else None,
                     round(pg, 1) if pg else None, shared, nb])
    sheet(wb, "Топ-15 позиции",
               ["сайт", "показов в топ-15", "показов у сайта всего", "доля топ-15",
                "кликов в топ-15", "ср. позиция Яндекс", "ср. позиция Google",
                f"из {N} запросов есть у своих", "сколько сайтов пересекается"],
               data, [30, 15, 16, 11, 13, 13, 13, 15, 15],
               {2: "#,##0", 3: "#,##0", 4: "0.0%", 5: "#,##0", 6: "0.0", 7: "0.0"})

    # --- Топ-15 запросы ---
    data = []
    for s in sites:
        for k, q in enumerate(top[s], 1):
            py, pg = agg_ya.get((s, q)), agg_go.get((s, q))
            others = [f"{o} {agg[(o, q)][2]:.0f}" for o in sites if o != s and (o, q) in agg]
            data.append([s, k, q, idx[s][q][0], idx[s][q][1],
                         round(py[2], 1) if py else None, round(pg[2], 1) if pg else None,
                         len(others), ", ".join(others)])
    sheet(wb, "Топ-15 запросы",
          ["сайт", "№", "запрос", "показы", "клики", "позиция Яндекс", "позиция Google",
           "своих сайтов рядом", "другие мои сайты по запросу (позиция)"],
          data, [28, 5, 44, 10, 8, 13, 13, 12, 90],
          {4: "#,##0", 5: "#,##0", 6: "0.0", 7: "0.0"})

    # --- Матрица топ-15 ---
    data = []
    for s in sites:
        data.append([s] + ["—" if o == s else (len(tset[s] & aset[o]) or None) for o in sites])
    ws = sheet(wb, "Матрица топ-15", ["из топ-15 строки есть у столбца"] + sites, data,
               [32] + [9] * len(sites), freeze="B2")
    for c in range(2, len(sites) + 2):
        ws.cell(1, c).alignment = Alignment(text_rotation=90, vertical="bottom", horizontal="center")
    ws.row_dimensions[1].height = 130

    data = []
    for s in sites:
        data.append([s] + ["—" if o == s else (len(tset[s] & tset[o]) or None) for o in sites])
    ws = sheet(wb, "Матрица голова в голову", ["в топ-15 обоих сайтов"] + sites, data,
               [32] + [9] * len(sites), freeze="B2")
    for c in range(2, len(sites) + 2):
        ws.cell(1, c).alignment = Alignment(text_rotation=90, vertical="bottom", horizontal="center")
    ws.row_dimensions[1].height = 130

    # --- Пересечение по всей семантике ---
    _, _, pairs = bo.pair_stats(idx)
    data = [[p["a"], p["b"], p["n_a"], p["n_b"], p["shared"], p["overlap"], p["jaccard"],
             p["imp_share_a"], p["imp_share_b"], p["imp_shared"]]
            for p in sorted(pairs, key=lambda x: -x["shared"])]
    sheet(wb, "Пары по всем запросам",
          ["сайт A", "сайт B", "запросов A", "запросов B", "общих запросов",
           "коэф. перекрытия", "Жаккар", "показов A на общих", "показов B на общих",
           "показов на общих"],
          data, [28, 28, 11, 11, 13, 13, 10, 14, 14, 14],
          {3: "#,##0", 4: "#,##0", 5: "#,##0", 6: "0.0%", 7: "0.0%",
           8: "0.0%", 9: "0.0%", 10: "#,##0"})

    # --- Спорные запросы ---
    cont = sorted(((sum(agg[(d, q)][0] for d in o), q, o)
                   for q, o in owners.items() if len(o) > 1), reverse=True)[:2000]
    data = []
    for imp, q, o in cont:
        ds = sorted(o, key=lambda d: -agg[(d, q)][0])
        data.append([q, len(ds), imp, sum(agg[(d, q)][1] for d in ds),
                     ds[0], round(agg[(ds[0], q)][2], 1),
                     ds[1], round(agg[(ds[1], q)][2], 1),
                     " | ".join(f"{d} {agg[(d, q)][0]}/{agg[(d, q)][1]}" for d in ds)])
    sheet(wb, "Спорные запросы",
          ["запрос", "сайтов", "показов", "кликов", "сайт 1", "позиция 1",
           "сайт 2", "позиция 2", "все сайты (показы/клики)"],
          data, [44, 8, 10, 9, 26, 10, 26, 10, 80],
          {3: "#,##0", 4: "#,##0", 6: "0.0", 8: "0.0"})

    wb.save(OUT)
    print("сохранено:", OUT)


if __name__ == "__main__":
    main()
