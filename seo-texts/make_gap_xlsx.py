# -*- coding: utf-8 -*-
"""group-gap-candidates.csv -> group-gap-candidates.xlsx

Колонка «значение» пишется как текст: иначе Excel в ru-локали превращает
«7,5» в дату 07.май, а «IP54» оставляет, и список едет.

Сводка посчитана значениями, не формулами: LibreOffice в этой среде не
поднимается, а формулу без пересчёта openpyxl отдаёт с пустым кешем.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROWS = list(csv.DictReader(open('group-gap-candidates.csv', encoding='utf-8-sig'), delimiter=';'))
COLS = ['тир', 'блок', 'фасет', 'ключ', 'значение', 'товаров', 'шорт-лист']
WIDTH = [7, 17, 27, 11, 34, 11, 11]

INK   = '1F2A2B'
HEAD  = '0C6B6B'
TIER  = {'1': 'FDEBD8', '2': 'E4F0EF', '3': None, '-': 'F0F0EE', '': None}
FONT  = 'Arial'
hair  = Side(style='thin', color='DCE2E0')

wb = Workbook()

# ------------------------------------------------------------- лист «Кандидаты»
ws = wb.active
ws.title = 'Кандидаты'
ws.append(COLS)
for c in ws[1]:
    c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=HEAD)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 26

for r in ROWS:
    ws.append([r['тир'], r['блок'], r['фасет'], r['ключ'], r['значение'],
               int(r['товаров']), r['шорт-лист']])
    row = ws.max_row
    fill = TIER.get(r['тир'])
    for i in range(1, len(COLS) + 1):
        c = ws.cell(row=row, column=i)
        c.font = Font(name=FONT, size=10, color=INK)
        c.border = Border(bottom=hair)
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=5).number_format = '@'      # текст, иначе 7,5 -> дата
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')

for i, w in enumerate(WIDTH, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{ws.max_row}'
LAST = ws.max_row

# ---------------------------------------------------------------- лист «Сводка»
sm = wb.create_sheet('Сводка')
sm.column_dimensions['A'].width = 46
for col, w in (('B', 13), ('C', 15), ('D', 46)):
    sm.column_dimensions[col].width = w

def put(row, vals, bold=False, size=10, color=INK):
    for i, v in enumerate(vals, 1):
        c = sm.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=size, bold=bold, color=color)
    return row + 1

R = put(1, ['Раздел «Воздушные компрессоры», выгрузка Битрикса'], bold=True, size=13)
R = put(R, ['20 046 активных публикуемых позиций. Порог группы — больше 10 товаров.'], color='5D6D6D')
R += 1
R = put(R, ['Тир', 'Групп', 'Позиций*', 'Что это'], bold=True)
for c in sm[R - 1]:
    c.fill = PatternFill('solid', fgColor='E4F0EF')

PLAN = [('1', 'делать сразу — дешёвые дырки с готовой семантикой'),
        ('2', 'серии, вложенные в бренд'),
        ('3', 'по остаточному принципу, спрос жиже'),
        ('-', 'порог прошло, делать не советую'),
        ('',  'не прошло отсев (мусорные значения)')]
first = R
for t, note in PLAN:
    sel = [r for r in ROWS if r['тир'] == t]
    sm.cell(row=R, column=1, value=t or 'без тира').font = Font(name=FONT, size=10, bold=True, color=INK)
    sm.cell(row=R, column=2, value=len(sel))
    sm.cell(row=R, column=3, value=sum(int(r['товаров']) for r in sel))
    sm.cell(row=R, column=4, value=note)
    for i in (2, 3):
        sm.cell(row=R, column=i).font = Font(name=FONT, size=10, color=INK)
        sm.cell(row=R, column=i).number_format = '#,##0'
    sm.cell(row=R, column=4).font = Font(name=FONT, size=10, color='5D6D6D')
    R += 1
sm.cell(row=R, column=1, value='Всего кандидатов').font = Font(name=FONT, size=10, bold=True, color=INK)
sm.cell(row=R, column=2, value=len(ROWS)).font = Font(name=FONT, size=10, bold=True, color=INK)
sm.cell(row=R, column=2).number_format = '#,##0'
R += 2
R = put(R, ['* Позиции считаются с пересечениями: один компрессор попадает сразу в несколько групп.'], color='5D6D6D')
R += 1
R = put(R, ['Колонка «значение» на листе «Кандидаты» — текстовая. Если переформатировать'], color='5D6D6D')
R = put(R, ['её в число, Excel в ru-локали превратит «7,5» обратно в дату 07.май.'], color='5D6D6D')
R += 1
R = put(R, ['Цифры — срез на момент выгрузки, не формулы: при ручной правке списка они не пересчитаются.'], color='5D6D6D')
R += 1
R = put(R, ['Источник: bitrix-export-full.tar.gz (дроп). Сборка: seo-texts/make_gap_xlsx.py'], color='8B9A99')

wb.save('group-gap-candidates.xlsx')
print('строк данных:', LAST - 1)
