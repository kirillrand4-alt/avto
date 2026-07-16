# -*- coding: utf-8 -*-
"""Коммерческое ранжирование акцепторов (v2, правки владельца):
1) B2B CTR-кривая: плоская в топ-5 (заявки на КП оставляют на многих сайтах),
   ступени на входе в топ-5/топ-10, обрыв на 2-й странице.
2) Бот-фильтр показов: выдачу постоянно сканируют (особенно «винтовые»);
   эффективные показы восстанавливаем из фактических кликов (implied = clicks/CTR(pos)).
3) Ценность страницы = медианный чек сегмента из prokompressor-прайса (price-map.json),
   а не ручные веса. value = прирост кликов * реалистичность * чек(млн ₽).
4) Доминирование выдачи: в B2B клиент кликает несколько сайтов из одного СЕРПа -
   сегменты, где >=2 наших сайтов в зоне 4-30, выгодно качать пакетом.
Вход: выгрузка all_sites + frog/urls-indexable-all.txt + frog/price-map.json
Выход: frog/acceptor-value.json + frog/acceptor-value.xlsx (5 листов)."""
import json, os, re, statistics
import openpyxl
from openpyxl.styles import Font

DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_IN = '/root/.claude/uploads/bcce55cd-293a-515c-9700-ae71a77daa5a/382f529d-all_sites_page_20260616_20260715.xlsx'

# --- B2B CTR-кривая ---
def ctr(p):
    p = max(1.0, p)
    if p < 2: return .120
    if p <= 3: return .090
    if p <= 5: return .075
    if p <= 7: return .050
    if p <= 10: return .035
    if p <= 15: return .012
    if p <= 20: return .007
    return .003

# реалистичность подъёма ссылками: с 1-й страницы проще, со 2-й - долго/дорого
def feasibility(p):
    if p <= 10: return 1.0
    if p <= 15: return 0.7
    if p <= 20: return 0.5
    return 0.3

# --- сегменты URL -> ключи прайса (первое совпадение) ---
URL_SEG = [
    ('gen_kislorod', 'Кислородные станции', r'kislorod|oxygen'),
    ('gen_azot', 'Азотные станции', r'azot|nitrogen|generatsiya-gazov'),
    ('mks', 'Модульные/блочные КС', r'modulnye-kompressornye|blochn|kompressornye-stantsii'),
    ('vd_buster', 'Дожимные/ВД', r'dozhim|booster|vysokogo-davleniya'),
    ('vint_diz', 'Дизельные/передвижные', r'dizel|peredvizh|shassi'),
    ('bezmasl', 'Безмасляные/центробежные', r'centrobezh|tsentrobezh|turbo|bezmasl|oil-free\.ru|spiraln|scroll'),
    ('vint_el', 'Винтовые', r'vintov|screw|kompressory-s-resiverom'),
    ('osush_ads', 'Осушители адсорбц.', r'adsorb'),
    ('osush_mix', 'Осушители', r'osushitel|refrizh'),
    ('filtry', 'Фильтры/запчасти/масла', r'filtr|separator|maslo|zapchast|remkomplekt|raskhodn|kartridzh'),
    ('resivery', 'Ресиверы', r'resiver|vozduhosborn|vozdukhosborn'),
    ('vozduhoduvki', 'Воздуходувки', r'vozduhoduv|vozdukhoduv'),
    ('porshn', 'Поршневые/бытовые', r'porshnev|remenn|pryamym-privod|koaksial|avtomobiln|garazh'),
]
def segment(url):
    u = url.lower()
    for key, name, rx in URL_SEG:
        if re.search(rx, u):
            return key, name
    return 'vint_el', 'Бренд/категория (смеш.)'   # брендовые сайты: ядро продаж - винтовые

# --- типы станций из списка владельца ---
STATION_TYPES = [
    ('Азотная станция', r'azot|nitrogen'),
    ('Кислородная станция', r'kislorod|oxygen'),
    ('Передвижная компрессорная станция', r'peredvizh|dizel|shassi'),
    ('Модульная/блочная станция', r'modul|blochn'),
    ('Дожимная (бустер)', r'dozhim|booster'),
    ('Компрессорная станция (КС общая)', r'stan(c|ts)i'),
]
ST_ANY = re.compile(r'azot|kislorod|nitrogen|oxygen|dozhim|booster|blochn|stan(c|ts)i|'
                    r'generator', re.I)
ST_NOT = re.compile(r'/blog/|/poleznoe/|/news/|/projects/|membrannye-osushiteli', re.I)


def load_prices():
    pm = json.load(open(os.path.join(DIR, 'price-map.json')))['segments']
    price = {k: v['median'] for k, v in pm.items()}
    price['osush_mix'] = round((price['osush_ads'] + price['osush_ref']) / 2)
    price['gen_kislorod'] = price['gen_azot']          # прокси: своих листов нет
    # МКС в прайсе - только контейнер; чек станции в сборе = ПБК + винтовик + осушитель + ресивер
    price['mks'] = price['mks'] + price['vint_el'] + price['osush_ads'] + price['resivery']
    return price


def norm_site(s):
    s = s.strip().lower().replace('sc-domain:', '')
    s = re.sub(r'^https?://', '', s).strip('/').replace('www.', '')
    return s


def load_stats():
    wb = openpyxl.load_workbook(XLSX_IN, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)
    best = {}   # дедуп sc-domain vs URL-property: (источник, сайт, url) -> max по показам
    for src, site, url, clicks, shows, _c, pos in rows:
        if not url or not shows:
            continue
        key = (src, norm_site(str(site)), str(url).strip().rstrip('/'))
        row = (int(clicks or 0), int(shows or 0), float(pos or 0))
        if key not in best or row[1] > best[key][1]:
            best[key] = row
    agg = {}
    for (src, site, url), (clicks, shows, pos) in best.items():
        a = agg.setdefault(url, dict(site=site, g_clicks=0, g_shows=0, y_clicks=0, y_shows=0,
                                     pos_num=0.0, pos_den=0))
        if src == 'Google':
            a['g_clicks'] += clicks; a['g_shows'] += shows
        else:
            a['y_clicks'] += clicks; a['y_shows'] += shows
        a['pos_num'] += pos * shows; a['pos_den'] += shows
    out = {}
    for url, a in agg.items():
        out[url] = dict(site=a['site'], shows=a['g_shows'] + a['y_shows'],
                        clicks=a['g_clicks'] + a['y_clicks'],
                        pos=round(a['pos_num'] / max(1, a['pos_den']), 1))
    return out


def bot_coefs(stats):
    """Медиана implied/shows по сегментам (страницы pos<=12, показы>=200):
    доля показов, подтверждённая фактическими кликами. Низкая = боты сканируют выдачу."""
    ratios = {}
    for url, s in stats.items():
        if s['pos'] <= 12 and s['shows'] >= 200:
            key, _ = segment(url)
            implied = s['clicks'] / ctr(s['pos'])
            ratios.setdefault(key, []).append(min(1.0, implied / s['shows']))
    return {k: round(min(1.0, max(0.15, statistics.median(v))), 3)
            for k, v in ratios.items() if len(v) >= 3}


def eff_shows(s, seg_key, coefs):
    """Эффективные показы: для топ-12 - из фактических кликов страницы,
    для 2-й страницы - сегментный бот-коэффициент."""
    if s['pos'] <= 12 and s['shows'] >= 200:
        implied = s['clicks'] / ctr(s['pos'])
        return min(s['shows'], max(implied, 0.15 * s['shows']))
    return s['shows'] * coefs.get(seg_key, 0.5)


def main():
    price = load_prices()
    stats = load_stats()
    coefs = bot_coefs(stats)
    indexable = set(l.strip().rstrip('/') for l in open(os.path.join(DIR, 'urls-indexable-all.txt')))

    cands = []
    for url, s in stats.items():
        if not (3.5 <= s['pos'] <= 30 and s['shows'] >= 30):
            continue
        if 'prokompressor.ru' not in s['site'] and url.rstrip('/') not in indexable:
            continue
        seg_key, seg_name = segment(url)
        kind = ('инфо' if re.search(r'/blog/|/poleznoe/|/news/|/stati/|instrukts|dokumentats', url)
                else 'главная' if url.rstrip('/').count('/') <= 2 else 'каталог')
        es = eff_shows(s, seg_key, coefs)
        gain = es * max(0.0, ctr(3) - ctr(s['pos']))
        chek = price.get(seg_key, price['vint_el'])
        value = gain * feasibility(s['pos']) * chek / 1e6
        cands.append(dict(url=url, site=s['site'], segment=seg_name, seg_key=seg_key,
                          kind=kind, price=chek, pos=s['pos'], shows=s['shows'],
                          eff_shows=round(es), clicks=s['clicks'],
                          gain=round(gain, 1), value=round(value, 1)))
    cands.sort(key=lambda c: -c['value'])

    # --- станции: инвентаризация + дыры ---
    st_pages = {}
    for url, s in stats.items():
        if ST_ANY.search(url) and not ST_NOT.search(url):
            st_pages[url.rstrip('/')] = dict(url=url, site=s['site'], pos=s['pos'],
                                             shows=s['shows'], clicks=s['clicks'], in_serp=True)
    for u in indexable:
        if ST_ANY.search(u) and not ST_NOT.search(u) and u not in st_pages \
                and u.split('://', 1)[-1].count('/') <= 3:
            st_pages[u] = dict(url=u, site=norm_site(u.split('/')[2]), pos=None,
                               shows=0, clicks=0, in_serp=False)
    st_list = sorted(st_pages.values(), key=lambda p: -(p['shows'] or 0))
    gaps = []
    for name, rx in STATION_TYPES:
        hits = [p for p in st_list if re.search(rx, p['url'], re.I)]
        gaps.append(dict(type=name, pages=len(hits),
                         best=max(hits, key=lambda p: p['shows'])['url'] if hits else '-',
                         shows=sum(p['shows'] for p in hits)))
    missing = ['Модульная кислородная станция', 'Дожимная кислородная станция',
               'Азотно-кислородная станция']

    # --- доминирование выдачи: сегменты с >=2 сайтами в зоне ---
    dom = {}
    for c in cands:
        if c['kind'] != 'каталог':
            continue
        d = dom.setdefault(c['seg_key'], {})
        cur = d.get(c['site'])
        if cur is None or c['value'] > cur['value']:
            d[c['site']] = c
    dom_rows = []
    for seg_key, sites in dom.items():
        if len(sites) < 2:
            continue
        pages = sorted(sites.values(), key=lambda x: x['pos'])
        dom_rows.append(dict(segment=pages[0]['segment'], seg_key=seg_key,
                             price=price.get(seg_key, 0), n_sites=len(pages),
                             sites=[(p['site'], p['pos'], p['eff_shows'], p['url']) for p in pages]))
    dom_rows.sort(key=lambda r: -(r['price'] * r['n_sites']))

    json.dump(dict(candidates=cands, stations=st_list, gaps=gaps, missing=missing,
                   bot_coefs=coefs, prices=price, domination=dom_rows),
              open(os.path.join(DIR, 'acceptor-value.json'), 'w'), ensure_ascii=False, indent=1)

    # --- xlsx ---
    wb = openpyxl.Workbook(); bold = Font(bold=True)
    ws = wb.active; ws.title = 'Рейтинг по ценности'
    ws.append(['#', 'URL', 'Сайт', 'Сегмент', 'Чек, ₽', 'Тип', 'Позиция', 'Показы/мес',
               'Показы без ботов', 'Клики/мес', 'Прирост кликов (топ-3..5)', 'Value (кл*чек, млн)'])
    for c in ws[1]: c.font = bold
    for i, c in enumerate(cands, 1):
        ws.append([i, c['url'], c['site'], c['segment'], c['price'], c['kind'], c['pos'],
                   c['shows'], c['eff_shows'], c['clicks'], c['gain'], c['value']])
    for col, w in zip('ABCDEFGHIJKL', (5, 72, 22, 24, 11, 9, 9, 11, 13, 10, 18, 14)):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet('Станции (всё что есть)')
    ws2.append(['URL', 'Сайт', 'Позиция', 'Показы/мес', 'Клики/мес', 'В выдаче'])
    for c in ws2[1]: c.font = bold
    for p in st_list:
        ws2.append([p['url'], p['site'], p['pos'] or '-', p['shows'], p['clicks'],
                    'да' if p['in_serp'] else 'нет (0 показов)'])
    for col, w in zip('ABCDEF', (80, 22, 9, 11, 10, 16)):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet('Дыры по типам станций')
    ws3.append(['Тип станции', 'Страниц', 'Показы суммарно', 'Лучшая страница'])
    for c in ws3[1]: c.font = bold
    for g in gaps:
        ws3.append([g['type'], g['pages'], g['shows'], g['best']])
    for m in missing:
        ws3.append([m, 0, 0, 'НЕТ СТРАНИЦЫ НИ НА ОДНОМ САЙТЕ'])
    for col, w in zip('ABCD', (36, 10, 15, 80)):
        ws3.column_dimensions[col].width = w

    ws4 = wb.create_sheet('Доминирование выдачи')
    ws4.append(['Сегмент', 'Чек, ₽', 'Сайтов в зоне 4-30', 'Сайт', 'Позиция', 'Показы без ботов', 'URL'])
    for c in ws4[1]: c.font = bold
    for r in dom_rows:
        for j, (site, pos, es, url) in enumerate(r['sites']):
            ws4.append([r['segment'] if j == 0 else '', r['price'] if j == 0 else '',
                        r['n_sites'] if j == 0 else '', site, pos, es, url])
    for col, w in zip('ABCDEFG', (24, 11, 10, 24, 9, 13, 72)):
        ws4.column_dimensions[col].width = w

    ws5 = wb.create_sheet('Чеки сегментов (из прайса)')
    ws5.append(['Сегмент (ключ)', 'Медианный чек, ₽', 'Бот-коэффициент показов'])
    for c in ws5[1]: c.font = bold
    for k in sorted(price, key=lambda k: -price[k]):
        ws5.append([k, price[k], coefs.get(k, '-')])
    for col, w in zip('ABC', (24, 16, 22)):
        ws5.column_dimensions[col].width = w
    wb.save(os.path.join(DIR, 'acceptor-value.xlsx'))

    # --- сводка ---
    print(f'Кандидатов: {len(cands)}')
    print('\nБот-коэффициенты (доля показов, подтверждённая кликами):')
    for k, v in sorted(coefs.items(), key=lambda kv: kv[1]):
        print(f'  {k:14} {v}')
    print('\nТОП-30 по value (клики*чек):')
    for c in cands[:30]:
        print(f"  {c['value']:>7.1f} | {c['segment'][:22]:22} чек {c['price']:>9,} | pos {c['pos']:>4} "
              f"| пок {c['shows']:>6}->{c['eff_shows']:>5} | {c['url'][:80]}")
    print('\nДоминирование (сегмент: сайтов в зоне):')
    for r in dom_rows[:10]:
        print(f"  {r['segment'][:24]:24} чек {r['price']:>9,} | {r['n_sites']} сайтов: " +
              ', '.join(f"{s}({p})" for s, p, _, _ in r['sites'][:6]))


if __name__ == '__main__':
    main()
