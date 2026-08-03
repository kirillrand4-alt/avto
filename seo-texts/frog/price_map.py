# -*- coding: utf-8 -*-
"""Медианный чек по сегментам из pricelist-slim.xlsx (313 листов, разные раскладки).
Логика: на каждом листе ищем колонки с заголовком 'цена' (НЕ 'закуп'), валюта из
заголовка (USD x курс, EUR x курс, руб как есть, CNY пропуск), собираем числа ниже.
Сегмент листа - по имени листа. Выход: frog/price-map.json + таблица в stdout."""
import json, os, re, statistics, warnings
warnings.filterwarnings('ignore')
import openpyxl

DIR = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(DIR, '..', 'pricelist-slim.xlsx')
USD, EUR = 77.4912, 88.5259

SHEET_SEG = [  # (сегмент, regex по имени листа) - первое совпадение
    ('gen_azot', r'ген\w*\s*\.?\s*азота|генераторы азота|оксимат|ЧЗМЭК'),
    ('gen_kislorod', r'кислород'),
    ('mks', r'ПБК|МКС|блок-контейнер'),
    ('vint_diz', r'диз|передвиж'),
    ('vd_buster', r'бустер|ВД|высокого davl|высокого давл|выс\. давл|40 бар'),
    ('bezmasl', r'безмасл|спир|турбо|центробеж|scroll'),
    ('vint_el', r'винт|Sollant винтовые|АТОМ'),
    ('osush_ads', r'адс|adsorb|осуш\w*\.?\s*адсорб'),
    ('osush_ref', r'осуш|реф'),
    ('filtry', r'фильтр|расходники|масло|сепаратор|циклон'),
    ('resivery', r'ресивер'),
    ('porshn', r'поршн|порш'),
    ('vozduhoduvki', r'воздуходув'),
    ('nd', r'НД'),
]
SKIP_SHEET = re.compile(r'главная|OLD|коэф|таблица|чиллер|градирн|частотн|редуктор|шкивы|'
                        r'пескостру|мед|blast|ULTRATECH|GMP', re.I)


def sheet_segment(name):
    for seg, rx in SHEET_SEG:
        if re.search(rx, name, re.I):
            return seg
    return None


def parse_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace('\xa0', '').replace(' ', '').replace(',', '.')
        if re.fullmatch(r'\d+(\.\d+)?', s):
            return float(s)
    return None


# правдоподобные ценовые коридоры сегментов, ₽ (для решения валюты и отсева мусора)
PLAUSIBLE = dict(vint_el=(100e3, 8e6), vint_diz=(800e3, 20e6), bezmasl=(300e3, 30e6),
                 porshn=(10e3, 1.5e6), vd_buster=(300e3, 20e6), osush_ref=(30e3, 3e6),
                 osush_ads=(80e3, 10e6), filtry=(1e3, 500e3), resivery=(20e3, 600e3),
                 gen_azot=(500e3, 60e6), gen_kislorod=(500e3, 60e6), mks=(500e3, 15e6),
                 vozduhoduvki=(150e3, 15e6), nd=(300e3, 15e6))


def extract(ws, seg):
    """[(валюта, цены...)] по колонкам с 'цена' в заголовке."""
    grid = list(ws.iter_rows(values_only=True))
    price_cols = {}          # col -> множитель валюты
    for ri, row in enumerate(grid[:40]):
        for ci, c in enumerate(row or ()):
            if not isinstance(c, str):
                continue
            h = c.lower()
            if 'цена' in h and 'закуп' not in h and 'cny' not in h:
                if 'usd' in h or '$' in h: mult = USD
                elif 'eur' in h or '€' in h: mult = EUR
                elif 'руб' in h or 'rub' in h: mult = 1.0
                else: mult = None       # решим по величине значений
                price_cols.setdefault(ci, (ri, mult))
    prices = []
    for ci, (hdr_ri, mult) in price_cols.items():
        vals = []
        for row in grid[hdr_ri + 1:]:
            if row and ci < len(row):
                n = parse_num(row[ci])
                if n and 100 <= n <= 80_000_000:
                    vals.append(n)
        if not vals:
            continue
        lo, hi = PLAUSIBLE.get(seg, (10e3, 60e6))
        if mult is None:                # валюту решаем по попаданию медианы в коридор сегмента
            med = statistics.median(vals)
            in_rub = lo <= med <= hi
            in_usd = lo <= med * USD <= hi
            if in_rub and not in_usd: mult = 1.0
            elif in_usd and not in_rub: mult = USD
            elif in_rub and in_usd:     # обе гипотезы в коридоре: ближе к центру коридора
                center = (lo * hi) ** .5
                mult = 1.0 if abs(med - center) <= abs(med * USD - center) else USD
            else:
                continue                 # мусорная колонка (артикулы, кВт и т.п.)
        prices += [v * mult for v in vals if lo <= v * mult <= hi]
    return prices


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    seg_prices, sheet_report = {}, []
    for name in wb.sheetnames:
        if SKIP_SHEET.search(name):
            continue
        seg = sheet_segment(name)
        if not seg:
            continue
        try:
            pr = extract(wb[name], seg)
        except Exception:
            pr = []
        if pr:
            seg_prices.setdefault(seg, []).extend(pr)
            sheet_report.append((name.strip(), seg, len(pr), statistics.median(pr)))
    out = {}
    for seg, pr in sorted(seg_prices.items()):
        pr.sort()
        out[seg] = dict(n=len(pr), p25=round(pr[len(pr) // 4]),
                        median=round(statistics.median(pr)),
                        p75=round(pr[len(pr) * 3 // 4]))
    json.dump(dict(segments=out, sheets=[dict(sheet=s, seg=g, n=n, median=round(m))
                                          for s, g, n, m in sheet_report]),
              open(os.path.join(DIR, 'price-map.json'), 'w'), ensure_ascii=False, indent=1)
    print(f"{'сегмент':14} {'n':>6} {'p25':>12} {'медиана':>12} {'p75':>12}")
    for seg, d in sorted(out.items(), key=lambda kv: -kv[1]['median']):
        print(f"{seg:14} {d['n']:>6} {d['p25']:>12,} {d['median']:>12,} {d['p75']:>12,}")


if __name__ == '__main__':
    main()
