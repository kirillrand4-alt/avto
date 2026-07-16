# -*- coding: utf-8 -*-
"""Каких моделей нет на сайте, но есть в прайсах поставщиков.
Сайт: битрикс-выгрузка (IE_NAME+IE_CODE). Прайсы: pricelist-slim.xlsx (колонки 'Модель').
Нормализация: верхний регистр, кириллица-двойники -> латиница, убрать пробелы/дефисы."""
import csv, json, os, re, warnings
warnings.filterwarnings('ignore')
import openpyxl

DIR = os.path.dirname(os.path.abspath(__file__))
CYR = str.maketrans('АВЕКМНОРСТУХаверкмнорстух', 'ABEKMHOPCTYXabepkmhopctyx')

def norm(s):
    s = str(s).translate(CYR).upper()
    s = re.sub(r'[\s\-–—_/.]', '', s)
    return s

def model_tokens(s):
    """Кандидаты-артикулы из строки: буквенно-цифровые с цифрой, длина 4-25."""
    out = set()
    for t in re.findall(r'[A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9\-./]{2,24}', str(s)):
        if re.search(r'\d', t) and re.search(r'[A-Za-zА-Яа-я]', t):
            n = norm(t)
            if 4 <= len(n) <= 22:
                out.add(n)
    return out

# --- сайт: модели из битрикса ---
SRC = '/tmp/claude-0/-home-user-avto/bcce55cd-293a-515c-9700-ae71a77daa5a/scratchpad/bitrix/export_file_2ov3pw9ju90trsle.csv'
csv.field_size_limit(10**9)
site = set()
r = csv.reader(open(SRC, encoding='utf-8-sig', errors='replace', newline=''), delimiter=';')
hdr = next(r); i_name, i_code = hdr.index('IE_NAME'), hdr.index('IE_CODE')
for row in r:
    if len(row) > max(i_name, i_code):
        site |= model_tokens(row[i_name])
        site |= model_tokens(row[i_code].replace('-', ''))
print(f'сайт: уникальных модельных токенов {len(site)}')

# --- прайсы: колонки 'Модель'/'наименование' ---
wb = openpyxl.load_workbook(os.path.join(DIR, '..', 'pricelist-slim.xlsx'), read_only=True, data_only=True)
SKIP = re.compile(r'главная|OLD|коэф|Таблица', re.I)
report, total_models, total_missing = [], 0, 0
for name in wb.sheetnames:
    if SKIP.search(name): continue
    ws = wb[name]
    grid = list(ws.iter_rows(values_only=True))
    mcol = None
    for ri, row in enumerate(grid[:25]):
        for ci, c in enumerate(row or ()):
            if isinstance(c, str) and re.fullmatch(r'\s*(модель|наименование)[\s\S]{0,20}', c.lower()):
                mcol = (ri, ci); break
        if mcol: break
    if not mcol: continue
    models = set()
    for row in grid[mcol[0]+1:]:
        if row and mcol[1] < len(row) and row[mcol[1]]:
            models |= model_tokens(row[mcol[1]])
    if len(models) < 3: continue
    missing = {m for m in models if m not in site}
    total_models += len(models); total_missing += len(missing)
    report.append((name.strip(), len(models), len(missing)))
report.sort(key=lambda x: -x[2])
print(f'прайсы: моделей {total_models}, НЕТ на сайте: {total_missing} ({total_missing/max(1,total_models):.0%})')
print('топ листов по дырам:')
for n, m, mi in report[:15]:
    print(f'  {mi:>4}/{m:<4} {n}')
json.dump([dict(sheet=n, models=m, missing=mi) for n, m, mi in report],
          open(os.path.join(DIR, 'gap-models-by-sheet.json'), 'w'), ensure_ascii=False, indent=1)
