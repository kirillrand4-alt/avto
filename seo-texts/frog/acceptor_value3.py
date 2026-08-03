# -*- coding: utf-8 -*-
"""Ранжирование акцепторов v3: раздельный расчёт по поисковикам на свежей выгрузке.

Почему v3 (проверка фильтров v2 на данных 04.07-02.08, см. DATA-QUALITY-FRESH.md):
1) Слитая CTR-кривая v2 ~= кривая Яндекса, а 76%% показов — Google с CTR в 3-5 раз
   ниже: implied-фильтр v2 резал «гугловость» страницы, а не ботов.
2) Позиции ПС расходятся: медиана |pos_G - pos_Y| = 4.5, у 19%% URL > 10.
3) В Яндекс-данных сидят рекламные utm/roistat/PAGEN-строки (CTR до 100%%) — мусор.
4) Накрутка кликов berg-compressor.com в Яндексе продолжается (топ-10 CTR 25-29%%).
5) Бот-показы Google точечные: подпись «десктоп>=90%% + CTR<0.5%% + показы>=300».

Модель: value = [sum по ПС: clean_shows_e * max(0, CTR_e(3-5) - CTR_e(pos_e)) * feas(pos_e)]
                * чек / 1e6 * бренд-буст
CTR-кривые фитятся на самой выгрузке (по бакетам позиций, без мусора и ботов).

Вход:  fresh-page CSV (Источник,Сайт,URL,Клики,Показы,CTR %%,Позиция — /stat/api/export-all)
       + fresh-device CSV (то же + Устройство) для бот-подписи Google.
Выход: acceptor-value-fresh.json + acceptor-value-fresh.xlsx.
Запуск: python3 acceptor_value3.py <fresh-page.csv> <fresh-device.csv>
"""
import csv, json, os, re, statistics, sys

import acceptor_value as v2   # сегменты, чек-каскад, бренд-буст, типы страниц

DIR = os.path.dirname(os.path.abspath(__file__))
PAGE_CSV = sys.argv[1] if len(sys.argv) > 1 else 'fresh-page-30d.csv'
DEV_CSV = sys.argv[2] if len(sys.argv) > 2 else 'fresh-device-30d.csv'

# сайты компрессорной сети (чужие проекты панели — meyer/usort/фотосепараторы — мимо)
NET = {'prokompressor.ru', 'nsk.prokompressor.ru', 'barnaul.prokompressor.ru',
       'vladivostok.prokompressor.ru', 'enger-air.ru', 'zif-kompressor.ru',
       'berg-compressor.com', 'berg-kompressor.ru', 'abac-kompressor.ru',
       'ekomak-kompressor.com', 'ekomak-compressor.com', 'crossair-compressor.ru',
       'dali-kompressor.ru', 'remeza-kompressor.ru', 'kraftmann-kompressor.com',
       'fini-compressor.com', 'ironmac-compressor.com', 'ac-kompressor.ru',
       'comaro-kompressor.ru', 'oil-free.ru', 'cmprg.ru', 'dizel-compressor.ru'}
CLICKBOT_Y = {'berg-compressor.com'}     # накрутка ПФ в Яндексе (свежее окно: топ-10 CTR 25-29%)

BUCKETS = [('3-5', 2.5, 5.5), ('6-10', 5.5, 10.5), ('11-15', 10.5, 15.5),
           ('16-20', 15.5, 20.5), ('21-30', 20.5, 30.5)]
def bucket(p):
    for name, lo, hi in BUCKETS:
        if lo <= p < hi:
            return name
    return None

def feas(p):
    if p <= 10: return 1.0
    if p <= 15: return 0.7
    if p <= 20: return 0.5
    if p <= 30: return 0.3
    return 0.0


def load_page(path):
    """CSV экспорта -> {(src, site, url): (clicks, shows, pos)}; дедуп sc-domain vs
    URL-property максимумом показов; мусорные строки с параметрами — вон."""
    best = {}
    with open(path, encoding='utf-8-sig') as f:
        rd = csv.reader(f); next(rd)
        for row in rd:
            src, site, url, cl, sh, _ctr, pos = row[:7]
            site = v2.norm_site(site)
            url = url.strip().rstrip('/')
            if site not in NET or '?' in url or not url.startswith('http'):
                continue
            try:
                rec = (int(float(cl or 0)), int(float(sh or 0)), float(pos or 0))
            except ValueError:
                continue
            if rec[1] <= 0:
                continue
            k = (src, site, url)
            if k not in best or rec[1] > best[k][1]:
                best[k] = rec
    return best


def load_desk_share(path):
    """Доля десктопа в показах Google по (site,url). Клики device-файла НЕ используем:
    экспорт устройств теряет ~3/4 кликов Google (проверено и на июньском, и на свежем)."""
    agg = {}
    with open(path, encoding='utf-8-sig') as f:
        rd = csv.reader(f); next(rd)
        for row in rd:
            src, site, url, dev, _cl, sh, _ctr, _pos = row[:8]
            if src != 'Google':
                continue
            k = (v2.norm_site(site), url.strip().rstrip('/'))
            d = agg.setdefault(k, [0, 0])
            sh = int(float(sh or 0))
            if (dev or '').upper() == 'DESKTOP':
                d[0] += sh
            d[1] += sh
    return {k: (d[0] / d[1], d[1]) for k, d in agg.items() if d[1] > 0}


def cut_clickbots(pages):
    """Яндекс-клики накрученных сайтов: кап CTR страницы 1.5x Яндекс-нормы позиции
    (норма — грубая стартовая, до фита кривых; клики этих сайтов в фит не идут)."""
    rough = {'3-5': .044, '6-10': .050, '11-15': .020, '16-20': .012, '21-30': .005}
    cut = 0
    for (src, site, url), (cl, sh, pos) in list(pages.items()):
        if src != 'Яндекс' or site not in CLICKBOT_Y:
            continue
        cap = round(sh * rough.get(bucket(max(pos, 3.0)) or '3-5', .04) * 1.5)
        if cl > cap:
            cut += cl - cap
            pages[(src, site, url)] = (cap, max(0, sh - (cl - cap)), pos)
    return cut


def fit_curves(pages, bot_flags):
    """Эмпирические CTR-бакеты по ПС: страницы показы>=50, без кликбот-Яндекса и
    без бот-подписных Google-страниц. Фолбэк — константы свежего окна."""
    default = {'Google': {'3-5': .0156, '6-10': .0095, '11-15': .0077, '16-20': .0045, '21-30': .0026},
               'Яндекс': {'3-5': .0450, '6-10': .0497, '11-15': .0196, '16-20': .0124, '21-30': .0046}}
    acc = {s: {b: [0, 0] for b, _, _ in BUCKETS} for s in default}
    for (src, site, url), (cl, sh, pos) in pages.items():
        if sh < 50 or (src == 'Яндекс' and site in CLICKBOT_Y) or (src, site, url) in bot_flags:
            continue
        b = bucket(max(pos, 3.0))
        if b and src in acc:
            acc[src][b][0] += cl; acc[src][b][1] += sh
    curves = {}
    for src, bs in acc.items():
        curves[src] = {}
        for b, (c, s) in bs.items():
            curves[src][b] = round(c / s, 4) if s >= 3000 else default[src][b]
    # цель «закрепление в топ-5»: у Яндекса топ-10 плоский — берём максимум из 3-5/6-10
    curves['Google']['target'] = curves['Google']['3-5']
    curves['Яндекс']['target'] = max(curves['Яндекс']['3-5'], curves['Яндекс']['6-10'])
    return curves


def ctr_of(curves, src, pos):
    b = bucket(pos)
    return curves[src][b] if b else 0.0


# --- v3.1: погранулярная кривая (проверка «топ-5 vs топ-7» на чистых данных окна).
# Яндекс: реальная ступень внутри топ-10 (поз.5-6 ~6.0%, поз.7 ~3.1%, поз.8-10 ~1.8%
#   — разница ×2-3, подтверждена объёмами ~6К показов на точку). Google: внутри
#   топ-10 почти плоско (1.0-1.7%, шум); его главная ступень — вход со 2-й страницы.
# Бакеты 3-5/6-10 это искажали: Яндекс-градиент недооценивался, Google-градиент
# переоценивался. Кривые сглажены вручную, чтобы не фитить шум малых ячеек.
CTR_FINE = {
    'Google': ((5.5, .014), (8.5, .011), (14.5, .0075), (20.5, .0046), (99, .0026)),
    'Яндекс': ((6.5, .060), (7.5, .031), (10.5, .018), (15.5, .016), (20.5, .0124), (99, .0046)),
}
TARGET_FINE = {'Google': .014, 'Яндекс': .060}   # CTR цели «закрепление в топ-5»


def ctr_fine(src, pos):
    for hi, v in CTR_FINE[src]:
        if pos < hi:
            return v
    return 0.0


def main():
    price = v2.load_prices()
    bx = json.load(open(os.path.join(DIR, 'bitrix-prices.json')))
    payload_prices = v2.load_payload_prices()
    indexable = set(l.strip().rstrip('/') for l in open(os.path.join(DIR, 'urls-indexable-all.txt')))

    pages = load_page(PAGE_CSV)
    desk = load_desk_share(DEV_CSV)
    clicks_cut = cut_clickbots(pages)

    # бот-подпись Google: десктоп>=90% (при базовых ~78% у живых сайтов), CTR<0.5%,
    # показы>=300, device-покрытие >=50% показов страницы
    bot_flags = {}
    for (src, site, url), (cl, sh, pos) in pages.items():
        if src != 'Google' or sh < 300:
            continue
        ds = desk.get((site, url))
        if ds and ds[1] >= sh * 0.5 and ds[0] >= 0.90 and cl / sh < 0.005:
            bot_flags[(src, site, url)] = round(ds[0], 3)

    curves = fit_curves(pages, bot_flags)

    # implied-валидация показов В КООРДИНАТАХ СВОЕЙ ПС (фикс главной ошибки v2):
    # сегментные медианы implied/shows для страниц, где клики валидируют показы
    seg_coef = {'Google': {}, 'Яндекс': {}}
    ratios = {'Google': {}, 'Яндекс': {}}
    for (src, site, url), (cl, sh, pos) in pages.items():
        if pos > 12 or sh < 200 or (src == 'Яндекс' and site in CLICKBOT_Y):
            continue
        c = ctr_fine(src, max(pos, 3.0))
        if c:
            seg, _ = v2.segment(url)
            ratios[src].setdefault(seg, []).append(min(1.0, (cl / c) / sh))
    for src, d in ratios.items():
        for seg, v in d.items():
            if len(v) >= 3:
                seg_coef[src][seg] = round(min(1.0, max(0.2, statistics.median(v))), 3)
    DEFAULT_COEF = {'Google': 0.8, 'Яндекс': 0.65}

    def clean_shows(src, site, url, cl, sh, pos, seg):
        """-> (показы без ботов, метод очистки)."""
        if (src, site, url) in bot_flags:      # точечные бот-страницы: только клики верим
            c = ctr_fine(src, max(pos, 3.0)) or .01
            return min(sh, max(cl / c, 0.05 * sh)), 'бот-подпись'
        if src == 'Яндекс' and site in CLICKBOT_Y:   # кликам не верим вовсе
            return sh * seg_coef[src].get(seg, DEFAULT_COEF[src]), 'кликбот-коэф'
        if pos <= 12 and sh >= 200:                  # клики страницы валидируют показы
            c = ctr_fine(src, max(pos, 3.0))
            return (min(sh, max(cl / c, 0.15 * sh)), 'implied') if c else (sh, '-')
        return sh * seg_coef[src].get(seg, DEFAULT_COEF[src]), 'сегм.коэф'

    # --- сборка кандидатов: (site,url) c вкладами обеих ПС ---
    by_url = {}
    for (src, site, url), rec in pages.items():
        by_url.setdefault((site, url), {})[src] = rec
    cands = []
    for (site, url), d in by_url.items():
        if 'prokompressor.ru' not in site and url not in indexable:
            continue
        seg_key, seg_name = v2.segment(url)
        contrib, tot_gain, parts = {}, 0.0, []
        for src, (cl, sh, pos) in d.items():
            if not (3.5 <= pos <= 30 and sh >= 30):
                continue
            es, method = clean_shows(src, site, url, cl, sh, pos, seg_key)
            g = es * max(0.0, TARGET_FINE[src] - ctr_fine(src, pos)) * feas(pos)
            contrib[src] = dict(pos=pos, shows=sh, clean=round(es), method=method,
                                clicks=cl, gain=round(g, 1))
            tot_gain += g
        if tot_gain <= 0:
            continue
        kind = ('инфо' if re.search(r'/blog/|/poleznoe/|/news/|/stati/|instrukts|dokumentats', url)
                else 'главная' if url.rstrip('/').count('/') <= 2 else 'каталог')
        chek, chek_src = v2.resolve_chek(url, site, seg_key, price, bx, payload_prices)
        boost, brand = v2.brand_boost(url, site)
        g_, y_ = d.get('Google'), d.get('Яндекс')
        cands.append(dict(
            url=url, site=site, segment=seg_name, seg_key=seg_key, kind=kind,
            price=chek, chek_src=chek_src, brand=brand,
            g_pos=g_ and round(g_[2], 1), y_pos=y_ and round(y_[2], 1),
            g_shows=g_ and g_[1], y_shows=y_ and y_[1],
            g_clicks=g_ and g_[0], y_clicks=y_ and y_[0],
            g_gain=round(contrib.get('Google', {}).get('gain', 0), 1),
            y_gain=round(contrib.get('Яндекс', {}).get('gain', 0), 1),
            g_clean=contrib.get('Google', {}).get('clean'),
            y_clean=contrib.get('Яндекс', {}).get('clean'),
            g_method=contrib.get('Google', {}).get('method', ''),
            y_method=contrib.get('Яндекс', {}).get('method', ''),
            bot_flag=(('Google', site, url) in bot_flags),
            gain=round(tot_gain, 1),
            value=round(tot_gain * chek / 1e6 * boost, 1)))
    cands.sort(key=lambda c: -c['value'])

    # --- доминирование: сегменты, где >=2 сайтов имеют каталог-кандидатов ---
    dom = {}
    for c in cands:
        if c['kind'] != 'каталог':
            continue
        d = dom.setdefault(c['seg_key'], {})
        if c['site'] not in d or c['value'] > d[c['site']]['value']:
            d[c['site']] = c
    dom_rows = []
    for seg_key, sites in dom.items():
        if len(sites) < 2:
            continue
        ps = sorted(sites.values(), key=lambda x: min(p for p in (x['g_pos'], x['y_pos']) if p))
        dom_rows.append(dict(segment=ps[0]['segment'], seg_key=seg_key,
                             price=price.get(seg_key, 0), n_sites=len(ps),
                             sites=[(p['site'], p['g_pos'], p['y_pos'], p['value'], p['url'])
                                    for p in ps]))
    dom_rows.sort(key=lambda r: -(r['price'] * r['n_sites']))

    meta = dict(model='v3.1 engine-split fine-grained', input=os.path.basename(PAGE_CSV),
                ctr_fine={k: list(v) for k, v in CTR_FINE.items()}, target=TARGET_FINE,
                clickbot_yandex_clicks_cut=clicks_cut,
                google_bot_pages=len(bot_flags),
                google_bot_shows=sum(pages[k][1] for k in bot_flags),
                curves=curves, seg_coef=seg_coef)
    json.dump(dict(meta=meta, candidates=cands, domination=dom_rows,
                   bot_pages=[dict(site=s, url=u, desk=v) for (_, s, u), v in bot_flags.items()]),
              open(os.path.join(DIR, 'acceptor-value-fresh.json'), 'w'),
              ensure_ascii=False, indent=1)

    # --- xlsx ---
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook(); bold = Font(bold=True)
    ws = wb.active; ws.title = 'Рейтинг v3 (по ПС раздельно)'
    ws.append(['#', 'URL', 'Сайт', 'Сегмент', 'Чек, ₽', 'Бренд', 'Тип',
               'G поз', 'Y поз', 'G показы', 'G без ботов', 'Y показы', 'Y без ботов',
               'G клики', 'Y клики', 'Прирост G', 'Прирост Y', 'Прирост кликов',
               'Value', 'Бот-флаг'])
    for c in ws[1]: c.font = bold
    for i, c in enumerate(cands, 1):
        ws.append([i, c['url'], c['site'], c['segment'], c['price'], c['brand'], c['kind'],
                   c['g_pos'] or '-', c['y_pos'] or '-', c['g_shows'] or 0, c['g_clean'] or 0,
                   c['y_shows'] or 0, c['y_clean'] or 0,
                   c['g_clicks'] or 0, c['y_clicks'] or 0, c['g_gain'], c['y_gain'],
                   c['gain'], c['value'], 'БОТЫ' if c['bot_flag'] else ''])
    for col, w in zip('ABCDEFGHIJKLMNOPQRST', (5, 66, 20, 22, 11, 9, 8, 7, 7, 9, 11, 9, 11, 8, 8, 10, 10, 12, 9, 9)):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet('CTR-кривые и коэф.')
    ws2.append(['ПС', 'бакет', 'CTR']); ws2['A1'].font = bold
    for src, bs in curves.items():
        for b, v in bs.items():
            ws2.append([src, b, v])
    ws2.append([]); ws2.append(['ПС', 'сегмент', 'implied-коэф показов'])
    for src, d in seg_coef.items():
        for seg, v in sorted(d.items()):
            ws2.append([src, seg, v])
    ws3 = wb.create_sheet('Бот-страницы Google')
    ws3.append(['Сайт', 'URL', 'Доля десктопа'])
    for c in ws3[1]: c.font = bold
    for (_, s, u), v in sorted(bot_flags.items(), key=lambda kv: -kv[1]):
        ws3.append([s, u, v])
    ws4 = wb.create_sheet('Доминирование')
    ws4.append(['Сегмент', 'Чек, ₽', 'Сайтов', 'Сайт', 'G поз', 'Y поз', 'Value', 'URL'])
    for c in ws4[1]: c.font = bold
    for r in dom_rows:
        for j, (site, gp, yp, val, url) in enumerate(r['sites']):
            ws4.append([r['segment'] if j == 0 else '', r['price'] if j == 0 else '',
                        r['n_sites'] if j == 0 else '', site, gp or '-', yp or '-', val, url])
    for col, w in zip('ABCDEFGH', (24, 11, 8, 22, 7, 7, 9, 66)):
        ws4.column_dimensions[col].width = w
    wb.save(os.path.join(DIR, 'acceptor-value-fresh.xlsx'))

    print(f"Кандидатов: {len(cands)} | срез Яндекс-кликов ({', '.join(CLICKBOT_Y)}): {clicks_cut}"
          f" | бот-страниц Google: {len(bot_flags)} ({meta['google_bot_shows']:,} показов)")
    print('\nCTR-кривые (фит на этой выгрузке):')
    for src, bs in curves.items():
        print(f"  {src:7}: " + '  '.join(f'{b}={v*100:.2f}%' for b, v in bs.items()))
    print('\nТОП-30 по value:')
    for c in cands[:30]:
        print(f"  {c['value']:>7.1f} | G {str(c['g_pos'] or '-'):>5} Y {str(c['y_pos'] or '-'):>5}"
              f" | пок G {c['g_shows'] or 0:>6} Y {c['y_shows'] or 0:>5}"
              f" | чек {c['price']:>10,} | {c['segment'][:20]:20} | {c['url'][:70]}"
              f"{' [БОТЫ]' if c['bot_flag'] else ''}")


if __name__ == '__main__':
    main()
