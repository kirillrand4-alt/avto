# -*- coding: utf-8 -*-
"""Пересборка отбора площадок: доперепроверка (dofollow глубже, живость, DDG-индексация)
+ Excel с исходными колонками, пересортированный по нашему приоритету."""
import json, os, re, subprocess, sys
from concurrent.futures import ThreadPoolExecutor
import openpyxl

DIR = os.path.dirname(os.path.abspath(__file__))
SRC = '/root/.claude/uploads/bcce55cd-293a-515c-9700-ae71a77daa5a/4b6a948c-______17.06.2026.xlsx'

SHORT = ['kineshemec.ru','gazetaingush.ru','gazetapervomaisk.ru','gazetapetrovka.ru','kz24.news','operativa.ru','samaraonline24.ru']
TOXIC_FAIL = {'bouw.ru':'редакционные размещения под займы/МФО','domocvet.com':'редакционные размещения под займы/МФО'}
WARN_MANUAL = {'arh112.ru':'проверить индексацию site: вручную','krasnodar.bz':'проверить индексацию site: вручную',
               'moscow-baku.ru':'проверить индексацию site: вручную','zheltaya.ru':'проверить индексацию site: вручную',
               'webferma.com':'антибот на curl - проверить глазами','new-sebastopol.com':'слабый dofollow - перепроверен ниже',
               'bobr.by':'нет dofollow в выборке','moiton.ru':'нет dofollow в выборке','oteplicah.com':'слабый dofollow',
               'ftimes.ru':'токсичность снята; трафик 80 - мёртвая?','gazetagavrilovka.ru':'токсичность снята; переплата за трафик'}

def curl(url, timeout=20):
    r = subprocess.run(['curl','-sL','--max-time',str(timeout),'-A',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0',url],
        capture_output=True, timeout=timeout+10)
    return r.stdout.decode('utf-8',errors='ignore')

def recheck(dom):
    ev={'domain':dom}
    try:
        page=curl(f'https://{dom}/')
        ev['alive']=len(page)>2000
        hrefs=re.findall(r'href="(https?://'+re.escape(dom)+r'[^"]{10,120})"',page)+['https://'+dom+h for h in re.findall(r'href="(/[^"]{10,120})"',page)]
        arts=[h for h in dict.fromkeys(hrefs) if re.search(r'/(news|stat|article|blog|post|20\d\d)/|\d{4,}',h)][:9]
        do=no=0
        for a in arts:
            html=curl(a,15)
            for mm in re.finditer(r'<a\s[^>]*href="(https?://[^"]+)"[^>]*>',html,re.I):
                if dom in mm.group(1): continue
                relm=re.search(r'rel="([^"]*)"',mm.group(0))
                if relm and re.search(r'nofollow|sponsored|ugc',relm.group(1)): no+=1
                else: do+=1
        ev['arts']=len(arts); ev['dofollow']=do; ev['nofollow']=no
        ddg=curl(f'https://html.duckduckgo.com/html/?q=site%3A{dom}',15)
        ev['ddg_results']=len(re.findall(r'result__a',ddg))
    except Exception as e:
        ev['err']=repr(e)[:80]
    return ev

print('доперепроверка шорт-листа...', flush=True)
with ThreadPoolExecutor(max_workers=7) as ex:
    rechecks={e['domain']:e for e in ex.map(recheck, SHORT)}
json.dump(rechecks, open(os.path.join(DIR,'shortlist-recheck.json'),'w'), ensure_ascii=False, indent=1)
for d,e in rechecks.items():
    print(f"  {d:22s} alive={e.get('alive')} статей={e.get('arts')} dofollow={e.get('dofollow')} nofollow={e.get('nofollow')} ddg={e.get('ddg_results')}", flush=True)

# порядок приоритета шорт-листа (мой named order с kineshemec первым)
ORDER = {d:i for i,d in enumerate(['kineshemec.ru','samaraonline24.ru','kz24.news','gazetapetrovka.ru','gazetapervomaisk.ru','operativa.ru','gazetaingush.ru'])}

def status_comment(dom, old_status, old_comment):
    r = rechecks.get(dom, {})
    extra = f" [перепроверка: dofollow {r.get('dofollow')}/{r.get('nofollow')} nofollow, ddg {r.get('ddg_results')}]" if r else ''
    if dom in TOXIC_FAIL:
        return 'НЕ БРАТЬ (фильтр)', f'токсичен: {TOXIC_FAIL[dom]}'
    if dom in ORDER:
        if r and r.get('dofollow',0)==0:
            return 'ПРОВЕРИТЬ (dofollow!)', f'шорт-лист, но доперепроверка не нашла dofollow{extra}'
        return f'БРАТЬ (шорт-лист #{ORDER[dom]+1})', (f'чист по всем проверкам{extra}; ' + (f'{old_comment}' if old_comment else '')).strip('; ')
    if dom in WARN_MANUAL:
        return 'РУЧНАЯ ПРОВЕРКА', WARN_MANUAL[dom]
    return old_status or '', old_comment or ''

def sort_key(dom):
    if dom in ORDER: return (0, ORDER[dom])
    if dom in WARN_MANUAL: return (1, 0)
    if dom in TOXIC_FAIL: return (2, 0)
    return (1, 5)

wb = openpyxl.load_workbook(SRC)
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0]); data = [list(r) for r in rows[1:] if any(r)]
    iD = hdr.index('Домен'); iS = hdr.index('Статус закупки'); iC = hdr.index('Комментарий') if 'Комментарий' in hdr else len(hdr)-1; iP = hdr.index('Приоритет')
    data.sort(key=lambda r: sort_key(str(r[iD]).strip().lower()))
    for i, r in enumerate(data, 1):
        dom = str(r[iD]).strip().lower()
        st, cm = status_comment(dom, r[iS], r[iC] if iC < len(r) else '')
        r[iP] = i; r[iS] = st
        if iC < len(r): r[iC] = cm
    for ri, r in enumerate(data, 2):
        for ci, v in enumerate(r, 1):
            ws.cell(row=ri, column=ci, value=v)
out = os.path.join(DIR, 'donors-prioritized.xlsx')
wb.save(out)
print('=>', out, flush=True)
