#!/usr/bin/env python3
"""Сборка MONTH-PLAN.md из трёх источников: солвер + съём с сервера + линзы.

    python3 render_month_plan.py
"""
from __future__ import annotations

import json
import re
import os

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'MONTH-PLAN.md')

SITE = {'enger': 'Enger', 'prokom': 'ProKompressor'}


def load(name):
    p = os.path.join(HERE, name)
    return json.load(open(p, encoding='utf-8')) if os.path.exists(p) else None


PRIO10 = ['kineshemec.ru', 'samaraonline24.ru', 'operativa.ru', 'moscow-baku.ru',
          'new-sebastopol.com', 'oteplicah.com', 'ftimes.ru', 'gazetagavrilovka.ru',
          'krasnodar.bz', 'arh112.ru']


def _wb(name):
    import openpyxl
    p = os.path.join(HERE, name)
    return openpyxl.load_workbook(p) if os.path.exists(p) else None


def scoring_cross():
    """Каждый из 10 доноров приоритета -> его судьба в свежем скоринге."""
    wb = _wb('donors-scored.xlsx')
    if not wb:
        return []
    sc, rj = wb['Скоринг'], wb['Отбраковка']
    hs = [str(c.value or '') for c in sc[1]]
    hr = [str(c.value or '') for c in rj[1]]
    S = {n: hs.index(n) for n in ['Домен', 'SCORE', 'Корзина', 'Трафик', 'ИКС', 'Статей']}
    R = {n: hr.index(n) for n in ['Домен', 'Причина', 'Трафик', 'ИКС']}

    good, bad = {}, {}
    for r in sc.iter_rows(min_row=2, values_only=True):
        good[str(r[S['Домен']] or '').lower()] = r
    for r in rj.iter_rows(min_row=2, values_only=True):
        bad[str(r[R['Домен']] or '').lower()] = r

    out = []
    for prio, dom in enumerate(PRIO10, 1):
        if dom in bad:
            r = bad[dom]
            prich = str(r[R['Причина']])
            # число размещений вытаскиваем только из причины «линкопомойка (N размещений)»,
            # иначе в колонку уезжает кусок текста про цену
            m = re.search(r'линкопомойка \((\d+)', prich)
            st = m.group(1) if m else '—'
            out.append((dom, prio, 'отбраковка: ' + prich, st, r[R['Трафик']], r[R['ИКС']]))
        elif dom in good:
            r = good[dom]
            out.append((dom, prio, f"прошёл, **{r[S['SCORE']]}** ({r[S['Корзина']]})",
                        r[S['Статей']], r[S['Трафик']], r[S['ИКС']]))
        else:
            out.append((dom, prio, '**отсутствует в базе Miralinks на 05.08**', '—', '—', '—'))
    out.sort(key=lambda x: x[1])
    return out


def replacement_shortlist(n=12):
    """Корзина A, отранжированная по «балл × измеренная релевантность».

    Релевантность считает donor_relevance.py по живым разделам площадки, а не по
    заявленной тематике — то есть ровно тем способом, который на списке 1-10 и
    показал расхождение.
    """
    wbs, wbr = _wb('donors-scored.xlsx'), _wb('donor-relevance.xlsx')
    if not (wbs and wbr):
        return []
    sc = wbs['Скоринг']
    hs = [str(c.value or '') for c in sc[1]]
    S = {n_: hs.index(n_) for n_ in ['Домен', 'SCORE', 'Трафик', 'Статей', 'Цена ₽']}
    score = {str(r[S['Домен']] or '').lower(): r for r in sc.iter_rows(min_row=2, values_only=True)}

    rl = wbr['Релевантность']
    hr = [str(c.value or '') for c in rl[1]]
    Rr = {n_: hr.index(n_) for n_ in ['Домен', 'Раздел размещения', 'AUDIENCE (раздел)',
                                      'Токс. %', 'МНОЖИТЕЛЬ']}
    rows = []
    for r in rl.iter_rows(min_row=2, values_only=True):
        dom = str(r[Rr['Домен']] or '').lower()
        s = score.get(dom)
        mult = r[Rr['МНОЖИТЕЛЬ']] or 0
        if not s or (s[S['SCORE']] or 0) < 72 or not mult:
            continue
        rows.append({
            'dom': dom, 'score': s[S['SCORE']], 'mult': mult,
            'total': round((s[S['SCORE']] or 0) * mult, 1),
            'razdel': r[Rr['Раздел размещения']], 'aud': r[Rr['AUDIENCE (раздел)']],
            'toks': r[Rr['Токс. %']], 'traf': s[S['Трафик']],
            'st': s[S['Статей']], 'price': s[S['Цена ₽']],
        })
    rows.sort(key=lambda x: -x['total'])
    return rows[:n]


def main() -> int:
    plan = load('month-plan.json')
    eyes = load('donor-eyes-raw.json') or {}
    lens = {x['donor']: x for x in (load('plan-lenses.json') or [])}

    L = []
    a = L.append
    a('# Месячная раскладка размещений (доноры 1-10 приоритета владельца)')
    a('')
    a('Задача владельца 05.08.2026: доноры 1-10 из `donors-prioritized.xlsx`, размещение')
    a('за один месяц, 1-3 ссылки в статье, доля Enger:ProKompressor = 40:60 в пользу')
    a('ProKompressor, каждый дилерский домен использован минимум один раз.')
    a('')
    a('Раскладку считает `plan_month.py` (перебор под жёсткие ограничения + независимая')
    a('валидация), доноров проверяет `donor_eyes.py` (браузер с сервера владельца,')
    a('дельфин-профили с мобильными прокси), подтверждают `plan_lenses.py` (три линзы')
    a('на разных моделях + судья).')
    a('')

    a('## Сводка')
    a('')
    a(f"| Показатель | Значение |")
    a('|---|---|')
    a(f"| Статей за месяц | {len(plan['rows'])} (по одной на донора) |")
    a(f"| Ссылок всего | {plan['links_total']} |")
    a(f"| На ядро (Enger + ProKompressor) | {plan['core_total']} |")
    a(f"| Enger / ProKompressor | {plan['enger']} / {plan['prokom']} = **{plan['ratio']}** |")
    a(f"| Дилерских ссылок | {plan['links_total'] - plan['core_total']} "
      f"(покрыты все {len(plan['dealers_covered'])}) |")
    a(f"| Проверка ограничений | {'нарушений нет' if not plan['validation'] else '; '.join(plan['validation'])} |")
    a('')

    a('## ⚠️ Заявленной тематике доноров верить нельзя (съём с сервера)')
    a('')
    a('Требование владельца: «тематике особо не верь у доноров, лучше перепроверь глазами')
    a('с сервера». Проверили — колонка тематики в выгрузке биржи проставлена вебмастером')
    a('и не верифицируется. Ниже — что площадка заявляет и что на ней реально.')
    a('')
    a('| # | Донор | Заявлено биржей | Разделы сайта по факту | Вывод линзы |')
    a('|---|---|---|---|---|')
    for r in sorted(plan['rows'], key=lambda x: x['prio']):
        dom = r['donor']
        e = eyes.get(dom, {})
        nav = ', '.join(x['text'] for x in e.get('nav', [])[:8]) or '— меню не снято'
        lz = lens.get(dom, {}).get('summary', {})
        verdict = lz.get('совпадение', '—')
        fact = lz.get('факт_тематика', '—')
        a(f"| {r['prio']} | {dom} | {r['profile']} | {nav} | {verdict}; факт: {fact} |")
    a('')

    a('## ⛔ Объективная сверка: список 1-10 не проходит собственный свежий скоринг')
    a('')
    a('Это проверка без моделей — сверка тех же десяти доменов с `donors-scored.xlsx`')
    a('(прогон базы Miralinks 14338 доноров, 05.08). Она весит больше вердикта линз,')
    a('потому что опирается на метрики биржи, а не на суждение.')
    a('')
    a('| # | Донор | Вердикт скоринга | Размещений | Трафик | ИКС |')
    a('|---|---|---|---|---|---|')
    for dom, prio, verdict, st, tr, iks in scoring_cross():
        a(f'| {prio} | {dom} | {verdict} | {st} | {tr} | {iks} |')
    a('')
    a('Ни один из десяти не попадает в корзину A. Порог отсечки по исходящим — 700')
    a('размещений через Miralinks; шестеро его превышают.')
    a('')
    a('**Две честные оговорки, без которых таблица выше вводит в заблуждение.**')
    a('')
    a('1. Порог 700 — грубое правило, заданное панелью из четырёх ИИ, а не владельцем.')
    a('   Оно не нормировано ни на возраст домена, ни на объём индекса. Кинешемец с 893')
    a('   размещениями и ручной июньской проверкой «dofollow 26/40» — не то же самое,')
    a('   что new-sebastopol с 5620. Ярлык «линкопомойка» здесь означает «превысил')
    a('   порог», а не «доказанная помойка».')
    a('2. Вердикт линз пришёл одинаковым по всем десяти («ОТЛОЖИТЬ ДОНОРА»), и часть')
    a('   этой одинаковости — дефект промпта: в линзе SEO-риска стояло «смотри придирчиво,')
    a('   ищи повод сказать нет», а линза тематики сравнивала площадку с заявленной')
    a('   тематикой, про которую уже было известно, что она неверна. Опираться следует на')
    a('   конкретные наблюдения линз, а не на их ярлык. Пример настоящего наблюдения:')
    a('   раздел «Экономика» у ftimes.ru по факту про ипотеку, вклады и ОСАГО.')
    a('')
    a('**Проверка этой оговорки.** Четырёх доноров (kineshemec, new-sebastopol,')
    a('gazetagavrilovka, arh112) первый прогон судил по неполному снимку — меню разделов')
    a('у них добралось только с третьей попытки. Перепрогнали их по полному снимку, другим')
    a('судьёй (gemini-3.6-flash вместо opus-4-8) и с разведёнными моделями линз. Вердикт')
    a('не изменился ни у одного; у kineshemec слегка вырос B2B-слой (НЕТ → СЛАБЫЙ) и место')
    a('размещения (3 → 4). То есть одинаковость вердиктов объясняется промптом лишь')
    a('частично — объективная сверка со скорингом указывает туда же.')
    a('')
    a('Самое предметное, что дал съём, — структура меню `arh112.ru`: поверх настоящего')
    a('портала (Служба-112, Росгвардия, Прокуратура, СК, УМВД) пришит второй блок рубрик')
    a('«Авто Мото, Бизнес и финансы, Недвижимость, Строительство, **Производство**, Работа,')
    a('Мебель и быт, Красота…». Это таксономия рубрик биржи, привинченная к сайту сводок')
    a('прокуратуры под платные статьи. Формально раздел «Производство» под нашу статью')
    a('там есть — но это силос платных размещений, а не редакция.')
    a('')

    a('## Раскладка')
    a('')
    a('| # | Донор | Страницы-акцепторы | Дилер | Ссылок | Место/релев. | Риск | Судья |')
    a('|---|---|---|---|---|---|---|---|')
    for r in sorted(plan['rows'], key=lambda x: x['prio']):
        core = '<br>'.join(
            f"№{c['n']} [{SITE[c['site']]}] {c['segment']}"
            + ('' if c['from_matrix'] else ' ⚠️вне матрицы')
            for c in r['core'])
        deal = ', '.join(x['key'] for x in r['dealers']) or '—'
        lz = lens.get(r['donor'], {}).get('summary', {})
        ready = ' *(статья готова)*' if r['ready'] else ''
        a(f"| {r['prio']} | {r['donor']}{ready} | {core} | {deal} | {r['links']} | "
          f"{lz.get('место', '—')}/{lz.get('релевантность', '—')} | {lz.get('риск', '—')} | "
          f"{lz.get('итог', '—')} |")
    a('')

    a('### Расход квоты акцепторов')
    a('')
    a('| № | Страница | Было ссылок | В этом месяце | Квота | Остаток |')
    a('|---|---|---|---|---|---|')
    seg = {c['n']: (c['segment'], c['site']) for r in plan['rows'] for c in r['core']}
    for n, u in sorted(plan['acceptor_usage'].items(), key=lambda x: int(x[0])):
        month = u['used_month'] - sum(1 for r in plan['rows'] if r['ready']
                                      for c in r['core'] if c['n'] == int(n))
        total = u['prior'] + month
        if not (month or u['prior']):
            continue
        s = seg.get(int(n), ('—', ''))
        a(f"| {n} | {s[0]} {('[' + SITE.get(s[1], '') + ']') if s[1] else ''} | {u['prior']} | "
          f"+{month} | {u['quota']} | {u['quota'] - total} |")
    a('')

    a('## Близкие альтернативы (в выбор не попали)')
    a('')
    a('Все проходят те же жёсткие ограничения и уступают победителю только по фит-баллу.')
    a('Малая разница = выбор в этом месте неустойчив, решает линза, а не формула.')
    a('')
    for alt in plan.get('alternatives', []):
        a(f"**A{alt['rank']}** — фит {alt['score']} ({alt['delta']:+}), отличий: {len(alt['diff'])}")
        a('')
        a('| Донор | В плане | В альтернативе |')
        a('|---|---|---|')
        for d in alt['diff']:
            was = ', '.join(f"№{x}" for x in d['was']['core']) + \
                  (' + ' + '/'.join(d['was']['dealers']) if d['was']['dealers'] else '')
            now = ', '.join(f"№{x}" for x in d['alt']['core']) + \
                  (' + ' + '/'.join(d['alt']['dealers']) if d['alt']['dealers'] else '')
            a(f"| {d['donor']} | {was} | {now} |")
        a('')

    if lens:
        a('## Что сказали линзы (подробно)')
        a('')
        for r in sorted(plan['rows'], key=lambda x: x['prio']):
            z = lens.get(r['donor'])
            if not z:
                continue
            s = z['summary']
            a(f"### {r['prio']}. {r['donor']}")
            a('')
            a(f"- **Тематика по факту:** {s['факт_тематика']} (совпадение с заявленным: "
              f"{s['совпадение']}, B2B-слой: {s['b2b_слой']})")
            a(f"- **Размещение:** место {s['место']}, релевантность {s['релевантность']} "
              f"→ {s['вердикт_релевантности']}")
            a(f"- **SEO-риск:** {s['риск']} → {s['решение_риска']}")
            a(f"- **Судья:** {s['итог']} (уверенность {s['уверенность']}) — {s['главное']}")
            a('')

    short = replacement_shortlist()
    if short:
        a('## Замена: шорт-лист из корзины A')
        a('')
        a('Ранжирование — «балл скоринга × измеренная релевантность». Релевантность здесь')
        a('считается по живым разделам площадки (`donor_relevance.py`), а не по заявленной')
        a('тематике: именно этот способ и вскрыл расхождение на списке 1-10.')
        a('')
        a('| Донор | Балл × релев. | Балл | Раздел размещения | AUD | Токс. % | Трафик | Размещений | Цена |')
        a('|---|---|---|---|---|---|---|---|---|')
        for x in short:
            a(f"| {x['dom']} | **{x['total']}** | {x['score']} | {x['razdel']} | {x['aud']} | "
              f"{x['toks']} | {x['traf']} | {x['st']} | {x['price']} |")
        a('')
        a('Это отраслевые площадки — кабельная промышленность, электроника, энергетика,')
        a('атомная отрасль, заводы. Статья про сжатый воздух там родной контент, и рубрика')
        a('под неё уже существует. Ровно того, чего нет ни у одного из десяти июньских.')
        a('')
        a('Предостережения по этому списку:')
        a('')
        a('- `dvobozrenie.ru` при балле 84 отдаёт **dofollow лишь в 29% размещений**')
        a('  (`DONOR-PAGE-AUDIT.md`) — брать только после проверки конкретной статьи.')
        a('- `ruscable.ru` (410 размещений) и `vpk.name` (471) — конвейер, хоть и профильный.')
        a('- Релевантность посчитана у 60 доменов из 191 в корзине A: список сузит не')
        a('  качество кандидатов, а недомер. Прогон остальных 131 расширит выбор.')
        a('')

    a('## Порядок работы')
    a('')
    a('1. Записать выбранные пары в `PLACEMENTS-LOG.md` **до** закупки.')
    a('2. Статьи готовить по `PLAYBOOK-GENERATION.md`, тон — `STYLE-GUIDE-GUEST.md`.')
    a('3. Темп: ≤2 новых донорских домена на наш сайт в неделю, ≤1 ссылка на один URL')
    a('   в две недели. Из-за этого №7 и №9 и №10, получающие в месяце по две ссылки,')
    a('   разносятся минимум на две недели.')
    a('4. Публикация на площадке — только после отдельного подтверждения владельца.')
    a('')

    open(OUT, 'w', encoding='utf-8').write('\n'.join(L) + '\n')
    print(f'-> {OUT} ({len(L)} строк)')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
