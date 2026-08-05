# -*- coding: utf-8 -*-
"""Релевантность донора: где живёт наш покупатель + токсичность + множитель к скорингу.

    python3 donor_relevance.py --from-scored donors-scored.xlsx --top 60 [--pages 24]
    python3 donor_relevance.py kineshemec.ru webferma.com
    python3 donor_relevance.py ... --keys keys-so-export.csv    # метод 2 (Keys.so/Serpstat)
    python3 donor_relevance.py --label sample.xlsx --top 40     # выборка под ручную разметку
    python3 donor_relevance.py --calibrate labeled.xlsx         # сверка модели с разметкой

РЕШЕНИЯ ВЛАДЕЛЬЦА 05.08.2026 (определяют всю конструкцию):
1. «Наша тема» = ГДЕ ЖИВЁТ НАШ ПОКУПАТЕЛЬ. Главная шкала - аудиторная (производство,
   стройка, СТО, агро, снабжение, энергетика), а не «пишет ли донор про компрессоры».
   Предметная близость (REL_TOPIC) - вторичный бонус, не условие.
2. Релевантность меряется НА УРОВНЕ РАЗДЕЛА, в котором вероятнее всего будет размещение,
   а не сайта целиком. Раздел определяется по имени (новости/статьи/бизнес/пресс-релизы)
   и по тому, где реально стоят внешние коммерческие ссылки.
3. Токсичность (казино/МФО/займы/ставки) - жёсткий порог: доля страниц с маркерами
   больше 1% = отказ, независимо от прочих баллов.
4. Итог применяется МНОЖИТЕЛЕМ к SCORE из score_donors.py: нерелевантный донор не
   спасается трафиком и трастом.
5. Модель калибруется ручной разметкой 30-40 доменов (--label / --calibrate).

Ловушка метода, найденная при калибровке: ОМОНИМЫ. «Ресивер» на доске объявлений -
спутниковый ТВ-ресивер (berkat.ru: 67 вхождений давали ложные 83 балла), «азотные» на
агросайте - удобрения, «рефрижератор» на транспортном - фура. Поэтому омонимичные
термины засчитываются только при наличии якоря предметной области на ТОЙ ЖЕ странице.
"""
import concurrent.futures as cf
import csv, glob, json, math, os, re, sys
import urllib.parse as up

import httpx

DIR = os.path.dirname(os.path.abspath(__file__))
SEO = os.path.dirname(DIR)
UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/126.0 Safari/537.36')

# ГЛАВНАЯ шкала: где живёт покупатель (решение владельца: п.1)
AUDIENCE_SEED = {
    'производств': 3, 'предприят': 3, 'завод': 3, 'цех': 3, 'промышленн': 3,
    'снабжен': 3, 'закупк': 2, 'модерниз': 2, 'оборудован': 2, 'подрядчик': 2,
    'строительн': 3, 'стройк': 2, 'стройплощадк': 3, 'прораб': 3, 'монтаж': 2,
    'спецтехник': 2, 'дорожн работ': 2,
    'автосервис': 3, 'шиномонтаж': 3, 'автомастерск': 3, 'автопарк': 2, 'грузов': 1,
    'фермер': 3, 'агропредприят': 3, 'элеватор': 3, 'зернохранилищ': 3, 'птицефабрик': 3,
    'теплиц': 2, 'хозяйств': 1, 'урожа': 1, 'посевн': 1,
    'деревообработ': 3, 'металлообработ': 3, 'сварк': 2, 'покрасочн камер': 3,
    'пищев производств': 3, 'энергетик': 2, 'котельн': 2, 'жкх': 1, 'логистик': 1,
    'бизнес': 1, 'инвестиц': 1, 'импортозамещ': 2, 'тендер': 2, 'госзакуп': 2,
}
# ВТОРИЧНАЯ шкала: наш предмет
TOPIC_SEED = {
    'компрессор': 3, 'компрессорн': 3, 'пневмат': 3, 'пневмосет': 3, 'пневмоинструмент': 3,
    'сжатого возд': 3, 'сжатый возд': 3, 'винтов': 2, 'поршнев': 2, 'ресивер': 3,
    'осушител': 3, 'точка рос': 3, 'адсорбц': 2, 'рефрижератор': 2, 'магистральн фильтр': 3,
    'азотн': 2, 'генератор азот': 3, 'кислородн': 2, 'генератор кислород': 3,
    'бустер': 2, 'дожимн': 3, 'безмасл': 3, 'воздуходувк': 2, 'конденсатоотвод': 3,
    'пневмоаудит': 3, 'отбойн молот': 2, 'пескостру': 2, 'краскопульт': 2, 'гайковёрт': 2,
}
# Токсичное соседство (решение владельца п.3: порог 1% страниц)
# ГРАНИЦЫ СЛОВ ОБЯЗАТЕЛЬНЫ: без них «комФОрт» ловится как «мфо», «слот» - в «слоток»,
# «ставк» - в «ставка ЦБ». Проверено на нашем же сайте: давал ложные 6.2% токсичности.
TOXIC_RX = re.compile(r'(\bказино|\bказик|\bбукмекер|\bставк\w* на спорт|\bбеттинг|\b1xbet|'
                      r'\bвулкан\b|\bмикрозайм|\bмикрокредит|\bзайм\w* на карт|\bмфо\b|'
                      r'\bденьги в долг|\bкредит онлайн|\bигров\w* автомат|\bбонус за регистрац|'
                      r'\bфрибет|\bпорно|\bэскорт|\bонлайн[- ]казино|\bслот\w*(?= |$)(?=[^а-я]*'
                      r'(игр|казино|бонус|автомат))|\bставки на спорт)', re.I)
# Омонимы: считаются только при якоре предметной области на той же странице
AMBIGUOUS = {'ресивер', 'азотн', 'кислородн', 'рефрижератор', 'компрессор', 'поршнев',
             'винтов', 'бустер'}
ANCHORS = ('сжат', 'пневмат', 'бар', 'м3/мин', 'давлен', 'осушител', 'компрессорн', 'пневмо')
# ДОСКИ ОБЪЯВЛЕНИЙ/МАРКЕТПЛЕЙСЫ - вторая ловушка (berkat.ru давал AUD 100 на
# перечислениях товаров в объявлениях). Аудиторные термины там - шум листингов,
# а не редакционный контент. Такие площадки штрафуются множителем 0.5.
CLASSIFIEDS_RX = re.compile(r'(подать объявлен|частн объявлен|доска объявлен|все объявлен|'
                            r'\bпродам\b|\bкуплю\b|\bсдам\b|\bсниму\b|цена договорн|'
                            r'торг уместен|объявлени\w* в категор|каталог товаров|'
                            r'добавить в корзину|товаров найдено)', re.I)
# Разделы, куда обычно попадает гостевое размещение
PLACEMENT_HINT = re.compile(r'(news|novosti|stat|article|blog|press|reklama|partner|promo|'
                            r'business|biznes|ekonomik|promyshl|stroit|agro|auto|avto)', re.I)
NON_CONTENT = re.compile(r'(login|search|cart|basket|contact|about|policy|feed|rss|tag|'
                         r'author|page/|\.(jpg|png|pdf|zip|css|js)$)', re.I)


def stem(w):
    return w[:6]


def fetch(url, client):
    try:
        r = client.get(url)
        return r.status_code, r.text
    except Exception:
        return 0, ''


def page_text(h):
    """Текст страницы; заголовки и title весомее (несут тему)."""
    h = re.sub(r'(?is)<(script|style)[^>]*>.*?</\1>', ' ', h)
    heads = ' '.join(re.findall(r'(?is)<h[1-3][^>]*>(.*?)</h[1-3]>', h))
    title = ' '.join(re.findall(r'(?is)<title[^>]*>(.*?)</title>', h))
    body = re.sub(r'<[^>]+>', ' ', h)
    return ' '.join([re.sub(r'<[^>]+>', ' ', heads)] * 3 + [title] * 3 + [body]).lower()


def seed_score(pages, seed, gate_ambiguous=False, scale=40.0):
    """Взвешенное покрытие словаря по страницам -> 0..100."""
    if not pages:
        return 0, {}
    total, ev = 0.0, {}
    for term, w in seed.items():
        cnt = pw = 0
        for p in pages:
            c = p.count(term)
            if not c:
                continue
            if gate_ambiguous and term in AMBIGUOUS and not any(a in p for a in ANCHORS):
                continue
            cnt += c; pw += 1
        if cnt:
            total += w * math.log1p(cnt) * (0.5 + 0.5 * pw / len(pages))
            ev[term] = cnt
    return min(100, round(100 * min(1.0, total / scale))), dict(
        sorted(ev.items(), key=lambda kv: -kv[1])[:8])


def crawl_sections(domain, pages_budget=24):
    """Главная -> разделы -> по 2-3 материала из каждого. Возвращает разделы с текстами."""
    with httpx.Client(follow_redirects=True, timeout=12, verify=False,
                      headers={'user-agent': UA}) as client:
        code, html = fetch(f'https://{domain}/', client)
        if code != 200 or not html:
            return None, code
        home = page_text(html)
        base = f'https://{domain}/'
        links = []
        for href in re.findall(r'href=["\']([^"\'#]+)["\']', html):
            u = up.urljoin(base, href)
            if domain.replace('www.', '') in u and not NON_CONTENT.search(u):
                links.append(u.split('?')[0])
        links = list(dict.fromkeys(links))
        # раздел = первый сегмент пути
        secs = {}
        for u in links:
            path = up.urlparse(u).path.strip('/')
            seg = path.split('/')[0] if path else ''
            if not seg or len(seg) > 40:
                continue
            secs.setdefault(seg, []).append(u)
        # приоритет разделам-кандидатам на размещение
        ranked = sorted(secs.items(),
                        key=lambda kv: (-bool(PLACEMENT_HINT.search(kv[0])), -len(kv[1])))[:6]
        per_sec = max(2, pages_budget // max(len(ranked), 1))
        out = {'__home__': [home]}
        jobs = []
        for seg, urls in ranked:
            for u in urls[:per_sec]:
                jobs.append((seg, u))
        with cf.ThreadPoolExecutor(max_workers=8) as ex:
            for (seg, u), (c, h) in zip(jobs, ex.map(lambda j: fetch(j[1], client), jobs)):
                if c == 200 and h:
                    out.setdefault(seg, []).append(page_text(h))
        return out, 200


def analyse(domain, pages_budget=24):
    secs, code = crawl_sections(domain, pages_budget)
    if secs is None:
        return {'domain': domain, 'status': f'http {code}',
                'verdict': 'НЕТ ДАННЫХ (нужен серверный обход)'}
    all_pages = [p for v in secs.values() for p in v]
    if len(all_pages) < 5:
        return {'domain': domain, 'status': 'мало страниц', 'pages': len(all_pages),
                'verdict': 'НЕТ ДАННЫХ (краул дал %d стр.)' % len(all_pages)}
    # токсичность - по всему собранному (решение п.3)
    # страница считается токсичной, если маркеры встречаются >=2 раз (одно упоминание
    # в новостной заметке - это не тематика площадки, а сюжет)
    tox_pages = [p for p in all_pages if len(TOXIC_RX.findall(p)) >= 2]
    tox_pct = round(100 * len(tox_pages) / len(all_pages), 1)
    tox_samples = []
    for p in tox_pages[:3]:
        m = TOXIC_RX.search(p)
        if m:
            tox_samples.append(p[max(0, m.start() - 40):m.start() + 40].strip())
    # по разделам
    sec_rows = []
    for seg, pages in secs.items():
        if len(pages) < 2 and seg != '__home__':
            continue
        aud, aud_ev = seed_score(pages, AUDIENCE_SEED, scale=40)
        top, top_ev = seed_score(pages, TOPIC_SEED, gate_ambiguous=True, scale=25)
        sec_rows.append({'section': seg, 'pages': len(pages), 'audience': aud, 'topic': top,
                         'placement_hint': bool(PLACEMENT_HINT.search(seg)),
                         'aud_ev': aud_ev, 'top_ev': top_ev})
    # раздел размещения: подсказка в имени -> лучший по аудитории; иначе лучший вообще
    cands = [s for s in sec_rows if s['placement_hint']] or \
            [s for s in sec_rows if s['section'] != '__home__'] or sec_rows
    place = max(cands, key=lambda s: (s['audience'], s['pages']))
    site_aud = max((s['audience'] for s in sec_rows), default=0)
    site_top = max((s['topic'] for s in sec_rows), default=0)
    r = {'domain': domain, 'status': 'ok', 'pages': len(all_pages),
         'sections': sec_rows, 'placement_section': place['section'],
         'place_audience': place['audience'], 'place_topic': place['topic'],
         'site_audience': site_aud, 'site_topic': site_top,
         'toxic_pct': tox_pct, 'toxic_pages': len(tox_pages), 'toxic_samples': tox_samples,
         'aud_ev': place['aud_ev'], 'top_ev': place['top_ev']}
    cls_pages = sum(1 for p in all_pages if len(CLASSIFIEDS_RX.findall(p)) >= 2)
    r['classifieds_pct'] = round(100 * cls_pages / len(all_pages))
    r['rel_mult'], r['verdict'] = relevance_multiplier(r)
    return r


def relevance_multiplier(r):
    """Решение п.4: релевантность - МНОЖИТЕЛЬ к SCORE (0 = отказ)."""
    if r.get('toxic_pct', 0) > 1.0:
        return 0.0, f"ОТКАЗ: токсичный контент на {r['toxic_pct']}% страниц (порог 1%)"
    aud, top = r.get('place_audience', 0), r.get('place_topic', 0)
    m = (1.00 if aud >= 60 else 0.90 if aud >= 45 else 0.75 if aud >= 30 else
         0.50 if aud >= 15 else 0.25)
    if top >= 45:
        m = min(1.25, m * 1.20)          # пишет и про наш предмет - редкость, поднимаем
    elif top >= 20:
        m = min(1.15, m * 1.08)
    cls = r.get('classifieds_pct', 0)
    if cls >= 30 and top < 45:           # доска объявлений: аудиторные слова - шум листингов
        m = min(m, 0.5)
    if cls >= 30 and top < 45:
        return round(m, 2), f'ДОСКА ОБЪЯВЛЕНИЙ ({cls}% страниц-листингов), множитель {round(m, 2)}'
    label = ('ЯДРО (аудитория и предмет наши)' if aud >= 45 and top >= 45 else
             'ЦЕЛЕВОЙ (покупатель живёт здесь)' if aud >= 60 else
             'БЛИЗКИЙ (аудитория частично)' if aud >= 45 else
             'СЛАБЫЙ (аудитория далека)' if aud >= 30 else
             'НЕРЕЛЕВАНТНЫЙ')
    return round(m, 2), f"{label}, множитель {round(m, 2)}"


# ---------- метод 2: запросы донора ----------

def keys_overlap(path):
    rows = []
    if path.lower().endswith(('.xlsx', '.xlsm')):
        import openpyxl
        data = list(openpyxl.load_workbook(path, read_only=True).active.iter_rows(values_only=True))
        hdr = [str(h or '').lower() for h in data[0]]
        rows = [dict(zip(hdr, r)) for r in data[1:]]
    else:
        with open(path, encoding='utf-8-sig', newline='') as f:
            rows = [{k.lower(): v for k, v in r.items()} for r in csv.DictReader(f)]
    agg = {}
    aud_terms = [t.split()[0] for t in AUDIENCE_SEED]
    top_terms = [t.split()[0] for t in TOPIC_SEED]
    for r in rows:
        dom = next((v for k, v in r.items() if 'домен' in k or 'domain' in k or 'сайт' in k), None)
        q = next((v for k, v in r.items() if 'запрос' in k or 'keyword' in k or 'phrase' in k), None)
        if not dom or not q:
            continue
        dom = re.sub(r'^https?://', '', str(dom)).replace('www.', '').strip('/').split('/')[0]
        a = agg.setdefault(dom, {'n': 0, 'aud': 0, 'top': 0, 'samples': []})
        a['n'] += 1
        ql = str(q).lower()
        if any(t in ql for t in aud_terms):
            a['aud'] += 1
            if len(a['samples']) < 5:
                a['samples'].append(str(q)[:60])
        if any(t in ql for t in top_terms):
            a['top'] += 1
    return {d: {'queries': v['n'], 'pct_aud': round(100 * v['aud'] / max(v['n'], 1)),
                'pct_top': round(100 * v['top'] / max(v['n'], 1)),
                'samples': '; '.join(v['samples'])} for d, v in agg.items()}


# ---------- калибровка (решение п.5) ----------

def make_label_sheet(res, path='relevance-label-sample.xlsx'):
    """Выборка под ручную разметку: разные корзины, вердикт скрыт от разметчика."""
    import openpyxl
    from openpyxl.styles import Font
    ok = [r for r in res if r.get('status') == 'ok']
    ok.sort(key=lambda r: -r.get('place_audience', 0))
    buckets = [ok[:12], ok[len(ok) // 2 - 6:len(ok) // 2 + 6], ok[-12:]]
    sample, seen = [], set()
    for b in buckets:
        for r in b:
            if r['domain'] not in seen:
                seen.add(r['domain']); sample.append(r)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Разметка'
    ws.append(['Домен', 'Раздел размещения', 'Твой вердикт (годен/нет/спорно)', 'Комментарий',
               'модель: AUDIENCE', 'модель: TOPIC', 'модель: множитель'])
    for c in ws[1]: c.font = Font(bold=True)
    for r in sample:
        ws.append([r['domain'], r.get('placement_section'), '', '',
                   r.get('place_audience'), r.get('place_topic'), r.get('rel_mult')])
    for col, wd in zip('ABCDEFG', (28, 22, 30, 40, 16, 14, 16)):
        ws.column_dimensions[col].width = wd
    wb.save(os.path.join(DIR, path))
    print(f'=> {path}: {len(sample)} доменов на ручную разметку (колонка C - твой вердикт)')


def calibrate(path):
    """Сверка модели с ручной разметкой: где расходимся."""
    import openpyxl
    ws = openpyxl.load_workbook(path, read_only=True).active
    rows = list(ws.iter_rows(values_only=True))[1:]
    tp = fp = fn = tn = 0
    mism = []
    for r in rows:
        dom, human, aud, mult = r[0], str(r[2] or '').strip().lower(), r[4], r[6]
        if not dom or not human:
            continue
        model_yes = (mult or 0) >= 0.75
        human_yes = human.startswith(('год', 'да', 'yes'))
        if model_yes and human_yes: tp += 1
        elif model_yes and not human_yes: fp += 1; mism.append((dom, 'модель ДА / ты НЕТ', aud, mult))
        elif not model_yes and human_yes: fn += 1; mism.append((dom, 'модель НЕТ / ты ДА', aud, mult))
        else: tn += 1
    n = tp + fp + fn + tn
    print(f'размечено {n}: совпадений {tp + tn} ({round(100 * (tp + tn) / max(n, 1))}%), '
          f'ложных ДА {fp}, ложных НЕТ {fn}')
    for d, why, aud, mult in mism:
        print(f'  {d:28} {why:22} AUDIENCE {aud} множитель {mult}')
    print('\nЧто крутить: много ложных ДА -> поднять порог аудитории в relevance_multiplier;'
          '\nмного ложных НЕТ -> расширить AUDIENCE_SEED терминами из этих доменов.')


def main():
    args = sys.argv[1:]
    if '--calibrate' in args:
        return calibrate(args[args.index('--calibrate') + 1])
    doms = [a for a in args if not a.startswith('--') and '.' in a
            and not a.endswith(('.xlsx', '.csv'))]
    if '--from-scored' in args:
        import openpyxl
        p = args[args.index('--from-scored') + 1]
        top = int(args[args.index('--top') + 1]) if '--top' in args else 50
        rows = list(openpyxl.load_workbook(p, read_only=True)['Скоринг'].iter_rows(values_only=True))
        doms = [str(r[1]) for r in rows[1:top + 1] if r[1]]
    if not doms:
        sys.exit('укажи домены или --from-scored donors-scored.xlsx --top N')
    budget = int(args[args.index('--pages') + 1]) if '--pages' in args else 24
    kso = keys_overlap(args[args.index('--keys') + 1]) if '--keys' in args else {}

    print(f'анализирую {len(doms)} доноров (бюджет {budget} стр./домен)...', flush=True)
    res = []
    with cf.ThreadPoolExecutor(max_workers=6) as ex:
        for r in ex.map(lambda d: analyse(d, budget), doms):
            k = kso.get(r['domain'].replace('www.', ''))
            if k:
                r.update(queries=k['queries'], pct_q_audience=k['pct_aud'],
                         pct_q_topic=k['pct_top'], query_samples=k['samples'])
            res.append(r)
            print(f"  {r['domain']:26} AUD {str(r.get('place_audience')):>4} | "
                  f"TOP {str(r.get('place_topic')):>4} | токс {str(r.get('toxic_pct')):>4}% | "
                  f"×{str(r.get('rel_mult')):>4} | {r.get('verdict')}", flush=True)

    json.dump(res, open(os.path.join(DIR, 'donor-relevance.json'), 'w'), ensure_ascii=False, indent=1)
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Релевантность'
    ws.append(['Домен', 'Стр.', 'Раздел размещения', 'AUDIENCE (раздел)', 'TOPIC (раздел)',
               'AUDIENCE (сайт)', 'Токс. %', 'Листингов %', 'МНОЖИТЕЛЬ', 'Вердикт',
               'Запросов', '% запросов аудитории', 'Примеры запросов',
               'Улики аудитории', 'Улики предмета'])
    for c in ws[1]: c.font = Font(bold=True)
    green = PatternFill('solid', fgColor='C6EFCE'); yellow = PatternFill('solid', fgColor='FFEB9C')
    red = PatternFill('solid', fgColor='FFC7CE')
    for r in sorted(res, key=lambda x: -(x.get('rel_mult') or 0)):
        ws.append([r['domain'], r.get('pages'), r.get('placement_section'),
                   r.get('place_audience'), r.get('place_topic'), r.get('site_audience'),
                   r.get('toxic_pct'), r.get('classifieds_pct'), r.get('rel_mult'), r.get('verdict'),
                   r.get('queries'), r.get('pct_q_audience'), r.get('query_samples'),
                   ', '.join(f'{k}:{v}' for k, v in (r.get('aud_ev') or {}).items()),
                   ', '.join(f'{k}:{v}' for k, v in (r.get('top_ev') or {}).items())])
        m = r.get('rel_mult')
        if m == 0: fill = red
        elif m and m >= 1.0: fill = green
        elif m and m >= 0.75: fill = yellow
        else: fill = None
        if fill:
            for c in ws[ws.max_row]: c.fill = fill
    for col, wd in zip('ABCDEFGHIJKLMNO', (26, 7, 20, 17, 15, 16, 9, 12, 11, 40, 10, 20, 40, 46, 40)):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = 'A2'
    ws2 = wb.create_sheet('Разделы')
    ws2.append(['Домен', 'Раздел', 'Страниц', 'AUDIENCE', 'TOPIC', 'Кандидат на размещение'])
    for c in ws2[1]: c.font = Font(bold=True)
    for r in res:
        for s in r.get('sections', []):
            ws2.append([r['domain'], s['section'], s['pages'], s['audience'], s['topic'],
                        'да' if s['placement_hint'] else ''])
    for col, wd in zip('ABCDEF', (26, 24, 9, 11, 9, 22)):
        ws2.column_dimensions[col].width = wd
    wb.save(os.path.join(DIR, 'donor-relevance.xlsx'))
    print('\n=> donor-relevance.xlsx / .json')
    if '--label' in args:
        make_label_sheet(res)


if __name__ == '__main__':
    main()
