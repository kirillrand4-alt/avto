# -*- coding: utf-8 -*-
"""Аудит РАЗМЕЩЕНИЙ донора: находит статьи с внешними ссылками и оценивает их качество.

    python3 audit_donor_pages.py домен1 домен2 ...            # разовый прогон
    python3 audit_donor_pages.py --from-scored donors-scored.xlsx --top 30
    python3 audit_donor_pages.py ... --ahrefs ur-export.csv    # домержить UR из Ahrefs

Идея владельца (05.08.2026): скачать статьи, из которых стоят внешние (рекламные)
ссылки, и посмотреть рейтинг САМИХ СТАТЕЙ. Если у 30-50%+ размещений рейтинг ненулевой -
домен хороший: значит его статьи живут (индексируются, перелинкованы, на них ссылаются),
а не лежат мёртвым грузом в архиве. У линкопомойки почти все размещения имеют UR=0.

Метрики считаются в два слоя:
  СЛОЙ 1 (свой краул, бесплатно): доля статей с внешними коммерческими ссылками
    (рекламность), rel-политика (dofollow/nofollow/sponsored), число внешних ссылок
    на статью, ВНУТРЕННИЕ ССЫЛКИ на статью (прокси UR: сирота в архиве или живая
    страница в рубриках), глубина от главной, свежесть.
  СЛОЙ 2 (Ahrefs, по выгрузке владельца): UR каждой статьи -> ДОЛЯ СТАТЕЙ С UR>=порога.
    Формат CSV/XLSX: колонки с URL и UR (имена ищутся регистронезависимо).

Выход: donor-pages-audit.json (сырьё) + donor-pages-audit.xlsx (2 листа: домены/статьи).
Ходит из песочницы обычным HTTPS; домены, отдающие 403/капчу, помечаются
`blocked` - их добивать через сервер (browser_probe с дельфин-профилями).
"""
import concurrent.futures as cf
import json, os, re, sys, time
import urllib.parse as up

import httpx

UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/126.0 Safari/537.36')
ART_RX = re.compile(r'/(news|article|articles|stat|stati|blog|post|posts|publikacii|'
                    r'materialy|press|reklama|partner)/|/20\d\d/|\d{5,}', re.I)
# домены, ссылка на которые НЕ считается рекламной
NEUTRAL = re.compile(r'(vk\.com|ok\.ru|t\.me|telegram|youtube|rutube|instagram|facebook|'
                     r'twitter|x\.com|dzen\.ru|yandex\.|google\.|gosuslugi|\.gov\.ru|'
                     r'kremlin\.ru|consultant\.ru|garant\.ru|wikipedia|\.pdf$|'
                     # Кнопки «поделиться» старого образца и служебные сервисы: на первом
                     # прогоне 300 доменов они дали 85 ложных «размещений» - digg,
                     # del.icio.us и stumbleupon встретились ровно по 18 раз каждый,
                     # то есть по одному разу на каждую статью шаблона.
                     r'digg\.com|del\.icio\.us|stumbleupon|reddit\.com|pinterest\.|'
                     r'tumblr\.com|livejournal\.|blogger\.com|addthis\.com|feedburner|'
                     r'host-tracker\.com|whatsapp\.|viber\.|max\.ru)', re.I)
COMMERCIAL_HINT = re.compile(r'(купить|цена|заказать|каталог|услуги|shop|store|catalog|price)', re.I)


def fetch(url, timeout=25, client=None, raw_ok=False):
    try:
        c = client or httpx.Client(follow_redirects=True, timeout=timeout,
                                   headers={'user-agent': UA}, verify=False)
        r = c.get(url)
        if raw_ok and r.content[:2] == b'\x1f\x8b':      # gzip-саймап
            import gzip
            return r.status_code, gzip.decompress(r.content).decode('utf-8', 'ignore')
        return r.status_code, r.text
    except Exception as e:
        return 0, f'__ERR__{e!r}'


def sitemap_urls(domain, client, limit=4000):
    """Собрать URL из sitemap (включая индексные sitemap-ы)."""
    seen, out = set(), []
    queue = [f'https://{domain}/sitemap.xml', f'https://{domain}/sitemap_index.xml',
             f'https://{domain}/sitemap-index.xml', f'https://{domain}/wp-sitemap.xml',
             f'https://{domain}/sitemap.xml.gz', f'https://{domain}/sitemap/index.xml']
    # robots.txt может указать нестандартный путь
    code, txt = fetch(f'https://{domain}/robots.txt', client=client)
    if code == 200:
        queue += re.findall(r'(?im)^\s*sitemap:\s*(\S+)', txt)[:5]
    while queue and len(out) < limit:
        sm = queue.pop(0)
        if sm in seen:
            continue
        seen.add(sm)
        code, txt = fetch(sm, client=client, raw_ok=sm.endswith('.gz'))
        if code != 200 or '<' not in txt:
            continue
        locs = re.findall(r'<loc>\s*([^<\s]+)\s*</loc>', txt)
        if '<sitemapindex' in txt[:2000].lower():
            queue += [l for l in locs if l not in seen][:12]
        else:
            out += locs
    return out[:limit]


def pick_articles(urls, domain, n=25):
    """Статьи-кандидаты: по маске, свежие (в конце sitemap обычно новее)."""
    arts = [u for u in urls if ART_RX.search(u) and domain in u]
    if not arts:
        arts = [u for u in urls if u.rstrip('/').count('/') >= 4]
    # берём с конца (свежие) + немного из середины для репрезентативности
    tail = arts[-n:] if len(arts) > n else arts
    mid = arts[len(arts) // 2: len(arts) // 2 + max(0, n - len(tail))]
    AD_HINT = re.compile(r'(reklama|partner|promo|advert|press-reliz|pressreliz|spec)', re.I)
    adish = [u for u in arts if AD_HINT.search(u)][:max(4, n // 3)]
    res, seen = [], set()
    for u in adish + list(reversed(tail)) + mid:
        if u not in seen:
            seen.add(u); res.append(u)
    return res[:n]


def norm_key(u):
    """Единый ключ URL: без схемы, www, query и хвостового слеша."""
    u = re.sub(r'^https?://', '', u.split('?')[0].split('#')[0])
    return u.replace('www.', '').rstrip('/').lower()


def analyse_article(url, domain, client):
    """Внешние ссылки статьи + rel-политика + признаки рекламности."""
    code, html = fetch(url, client=client)
    rec = {'url': url, 'http': code, 'ext_links': [], 'ext_dofollow': 0,
           'ext_nofollow': 0, 'sponsored': 0, 'commercial': 0, 'chars': 0,
           'has_ad_label': False}
    if code != 200 or html.startswith('__ERR__'):
        rec['error'] = html[:80] if html.startswith('__ERR__') else f'http {code}'
        return rec
    body = re.sub(r'(?is)<(script|style|nav|footer|header)[^>]*>.*?</\1>', ' ', html)
    rec['chars'] = len(re.sub(r'<[^>]+>', ' ', body))
    rec['has_ad_label'] = bool(re.search(r'на правах рекламы|партнёрск|партнерск|erid|'
                                         r'рекламодател|спонсорск', html, re.I))
    for m in re.finditer(r'<a\s([^>]*)href=["\']([^"\']+)["\']([^>]*)>(.*?)</a>', body, re.I | re.S):
        pre, href, post, anchor = m.group(1), m.group(2), m.group(3), m.group(4)
        if not href.startswith('http'):
            continue
        try:
            host = up.urlparse(href).netloc.replace('www.', '')
        except ValueError:
            continue        # «http://[…» и прочие кривые href: urlparse кидает ValueError
        if not host or domain.replace('www.', '') in host:
            continue
        if NEUTRAL.search(href) or NEUTRAL.search(host):
            continue
        rel = (re.search(r'rel=["\']([^"\']*)["\']', pre + post) or [None, ''])[1].lower()
        anchor_txt = re.sub(r'<[^>]+>', ' ', anchor).strip()[:60]
        is_sp = 'sponsored' in rel or 'ugc' in rel
        is_nf = 'nofollow' in rel
        rec['ext_links'].append({'host': host, 'href': href[:150], 'rel': rel,
                                 'anchor': anchor_txt})
        rec['sponsored'] += is_sp
        rec['ext_nofollow'] += (is_nf and not is_sp)
        rec['ext_dofollow'] += (not is_nf and not is_sp)
        if COMMERCIAL_HINT.search(anchor_txt) or COMMERCIAL_HINT.search(href):
            rec['commercial'] += 1
    return rec


def homepage_urls(domain, client, pages=10):
    """Фолбэк без sitemap: ссылки с главной и рубрик."""
    out = []
    code, html = fetch(f'https://{domain}/', client=client)
    if code != 200:
        return out
    def links(h, base):
        return [up.urljoin(base, x) for x in re.findall(r'href=["\']([^"\'#]+)["\']', h)]
    base = f'https://{domain}/'
    first = [u for u in links(html, base) if domain in u]
    out += first
    secs = [u for u in first if not ART_RX.search(u) and u.rstrip('/').count('/') <= 4]
    for sec in (secs or first)[:pages]:
        c, h = fetch(sec, client=client)
        if c == 200:
            out += [u for u in links(h, sec) if domain in u]
    return list(dict.fromkeys(out))


def internal_links_map(domain, client, pages=12):
    """Сколько внутренних ссылок ведёт на статьи (прокси UR: сирота или живая страница)."""
    counts = {}
    code, html = fetch(f'https://{domain}/', client=client)
    if code != 200:
        return counts
    sections = re.findall(r'href=["\'](/[^"\'#?]{3,60}/)["\']', html)
    targets = [f'https://{domain}/'] + [f'https://{domain}{s}' for s in
                                        list(dict.fromkeys(sections))[:pages]]
    for t in targets:
        c, h = fetch(t, client=client)
        if c != 200:
            continue
        for href in re.findall(r'href=["\']([^"\'#]+)["\']', h):
            full = up.urljoin(t, href)
            if domain.replace('www.', '') in full:
                k = norm_key(full)
                counts[k] = counts.get(k, 0) + 1
    return counts


def audit_domain(domain, n_articles=25):
    t0 = time.time()
    with httpx.Client(follow_redirects=True, timeout=25,
                      headers={'user-agent': UA}, verify=False) as client:
        code, _ = fetch(f'https://{domain}/', client=client)
        if code in (403, 429) or code == 0:
            return {'domain': domain, 'status': 'blocked', 'http': code,
                    'hint': 'добить через сервер: browser_probe + дельфин-профиль'}
        urls = sitemap_urls(domain, client)
        arts = pick_articles(urls, domain, n=n_articles)
        if not arts:                      # фолбэк: обход главной и рубрик
            urls = homepage_urls(domain, client)
            arts = pick_articles(urls, domain, n=n_articles)
        if not arts:
            return {'domain': domain, 'status': 'no-articles', 'sitemap_urls': len(urls)}
        inner = internal_links_map(domain, client)
        recs = []
        with cf.ThreadPoolExecutor(max_workers=5) as ex:
            for r in ex.map(lambda u: analyse_article(u, domain, client), arts):
                r['internal_links'] = inner.get(norm_key(r['url']), 0)
                recs.append(r)
    ok = [r for r in recs if r.get('http') == 200]
    # сайтовый шаблон: хост, встречающийся в >60% статей - это виджет/счётчик/партнёр
    # в футере, а не рекламное размещение. Вычитаем такие из статистики (урок 05.08).
    host_freq = {}
    for r in ok:
        for h in {l['host'] for l in r['ext_links']}:
            host_freq[h] = host_freq.get(h, 0) + 1
    # Порог снижен с 0.6 до 0.3 (прогон 18.08). Шестьдесят процентов пропускали самый
    # частый вид ложного «размещения» - ссылку на источник новости: belta.by стоит на
    # каждой перепечатке с анкором «БЕЛТА» или «Ссылка на оригинал», но перепечатки
    # составляют не всю ленту, а её часть, и до 0.6 хост не дотягивал. Треть статей
    # с одним и тем же внешним хостом - это уже шаблон, а не покупка.
    sitewide = {h for h, n in host_freq.items() if n >= max(3, 0.3 * max(len(ok), 1))}
    for r in ok:
        r['sitewide_hosts'] = sorted(sitewide & {l['host'] for l in r['ext_links']})
        keep = [l for l in r['ext_links'] if l['host'] not in sitewide]
        r['ext_links'] = keep
        r['ext_dofollow'] = sum(1 for l in keep if 'nofollow' not in l['rel'] and 'sponsored' not in l['rel'])
        r['ext_nofollow'] = sum(1 for l in keep if 'nofollow' in l['rel'] and 'sponsored' not in l['rel'])
        r['sponsored'] = sum(1 for l in keep if 'sponsored' in l['rel'] or 'ugc' in l['rel'])
    ad = [r for r in ok if r['ext_dofollow'] or r['ext_nofollow'] or r['sponsored']]
    ad_do = [r for r in ad if r['ext_dofollow']]
    orphan = [r for r in ad if r.get('internal_links', 0) == 0]
    return {
        'domain': domain, 'status': 'ok', 'sitemap_urls': len(urls),
        'checked': len(ok), 'with_ext_links': len(ad), 'with_dofollow': len(ad_do),
        'pct_ad': round(100 * len(ad) / max(len(ok), 1)),
        'pct_dofollow_of_ad': round(100 * len(ad_do) / max(len(ad), 1)),
        'pct_not_in_listings': round(100 * len(orphan) / max(len(ad), 1)),
        'avg_ext_per_ad': round(sum(len(r['ext_links']) for r in ad) / max(len(ad), 1), 1),
        'ad_labeled': sum(1 for r in ad if r['has_ad_label']),
        'sitewide_hosts': sorted(sitewide)[:8],
        'sponsored_links': sum(r['sponsored'] for r in ok),
        'top_ext_hosts': _top_hosts(ok),
        'seconds': round(time.time() - t0),
        'articles': recs,
    }


def _top_hosts(recs):
    c = {}
    for r in recs:
        for l in r['ext_links']:
            c[l['host']] = c.get(l['host'], 0) + 1
    return sorted(c.items(), key=lambda kv: -kv[1])[:10]


def merge_ahrefs(result, path, ur_threshold=5):
    """Домержить UR статей из выгрузки Ahrefs -> доля статей с рейтингом (метрика владельца)."""
    ur = {}
    if path.lower().endswith(('.xlsx', '.xlsm')):
        import openpyxl
        ws = openpyxl.load_workbook(path, read_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h or '').lower() for h in rows[0]]
        iu = next((i for i, h in enumerate(hdr) if 'url' in h), 0)
        ir = next((i for i, h in enumerate(hdr) if h.strip() in ('ur', 'url rating', 'рейтинг')), None)
        for r in rows[1:]:
            if r[iu] and ir is not None and r[ir] is not None:
                ur[str(r[iu]).split('?')[0].rstrip('/')] = float(r[ir])
    else:
        import csv
        with open(path, encoding='utf-8-sig', newline='') as f:
            rd = csv.DictReader(f)
            for row in rd:
                k = next((v for kk, v in row.items() if 'url' in kk.lower()), None)
                v = next((v for kk, v in row.items()
                          if kk.lower().strip() in ('ur', 'url rating', 'рейтинг')), None)
                if k and v not in (None, ''):
                    ur[k.split('?')[0].rstrip('/')] = float(v)
    for d in result:
        arts = [a for a in d.get('articles', []) if a.get('ext_dofollow')]
        vals = [ur.get(a['url'].split('?')[0].rstrip('/')) for a in arts]
        vals = [v for v in vals if v is not None]
        for a in arts:
            a['ur'] = ur.get(a['url'].split('?')[0].rstrip('/'))
        if vals:
            d['ur_known'] = len(vals)
            d['ur_median'] = sorted(vals)[len(vals) // 2]
            d['pct_ur_positive'] = round(100 * sum(1 for v in vals if v > 0) / len(vals))
            d['pct_ur_threshold'] = round(100 * sum(1 for v in vals if v >= ur_threshold) / len(vals))
    return result


def verdict(d):
    """Итог по домену в терминах владельца: доля живых размещений."""
    if d.get('status') != 'ok':
        return d.get('status', '?')
    if d.get('pct_ur_threshold') is not None:
        p = d['pct_ur_threshold']
        return ('ОТЛИЧНО (UR-доля %d%%)' % p if p >= 50 else
                'ХОРОШО (UR-доля %d%%)' % p if p >= 30 else
                'СЛАБО (UR-доля %d%%)' % p)
    # без Ahrefs судим по своим прокси
    if d['with_dofollow'] == 0:
        return 'ОТКАЗ: dofollow-размещений не найдено'
    if d['pct_dofollow_of_ad'] < 40:
        return 'СЛАБО: dofollow только у %d%% размещений' % d['pct_dofollow_of_ad']
    if d['avg_ext_per_ad'] > 4:
        return 'РИСК: в среднем %.1f внешних ссылок на статью' % d['avg_ext_per_ad']
    if d['with_dofollow'] < 3:
        return 'МАЛО ДАННЫХ: найдено %d dofollow-размещений (нужен Ahrefs)' % d['with_dofollow']
    return 'ГОДЕН по слою 1 (dofollow %d%%) - UR подтвердить Ahrefs' % d['pct_dofollow_of_ad']


def main():
    args = sys.argv[1:]
    doms = [a for a in args if not a.startswith('--') and '.' in a and not a.endswith(('.xlsx', '.csv'))]
    if '--from-scored' in args:
        import openpyxl
        path = args[args.index('--from-scored') + 1]
        top = int(args[args.index('--top') + 1]) if '--top' in args else 30
        ws = openpyxl.load_workbook(path, read_only=True)['Скоринг']
        rows = list(ws.iter_rows(values_only=True))
        doms = [str(r[1]) for r in rows[1:top + 1] if r[1]]
    if not doms:
        sys.exit('укажи домены или --from-scored donors-scored.xlsx --top N')
    n_art = int(args[args.index('--articles') + 1]) if '--articles' in args else 25
    print(f'аудит {len(doms)} доменов по {n_art} статей...', flush=True)
    # ВОЗОБНОВЛЯЕМОСТЬ. Прогон на сотне доменов идёт десятки минут, а песочница
    # перезапускается по нескольку раз в час - результат, сложенный только в память
    # до финального json.dump, при рестарте пропадает целиком (тот же урок, что с
    # потерянной выгрузкой биржи). Поэтому каждый домен дописывается строкой в jsonl
    # с fsync сразу, а повторный запуск пропускает уже собранное.
    CKPT = 'donor-pages-audit.jsonl'
    res, done = [], set()
    if os.path.exists(CKPT):
        for line in open(CKPT, encoding='utf-8'):
            line = line.strip()
            if not line:
                continue
            try:
                d = json.loads(line)
            except ValueError:
                continue
            if d.get('domain') not in done:
                done.add(d['domain'])
                res.append(d)
        if done:
            print(f'из чекпойнта поднято доменов: {len(done)}', flush=True)
    doms = [d for d in doms if d not in done]
    print(f'осталось собрать: {len(doms)}', flush=True)

    ck = open(CKPT, 'a', encoding='utf-8')
    # Воркеры разводят РАЗНЫЕ домены, внутри домена страницы качаются по очереди, поэтому
    # рост параллельности не бьёт по одной площадке. На 160 доменах доля отказов вышла 1%
    # (2 blocked), так что 4 потока были перестраховкой - для тысячи доменов их мало.
    nw = int(args[args.index('--workers') + 1]) if '--workers' in args else 8

    def safe(domain):
        """Падение одного домена не должно ронять прогон.

        Проверено дорогой ценой: на 220-м домене `urlparse` встретил href вида
        «http://[…» и бросил ValueError. Через ex.map исключение всплыло в цикл,
        цикл оборвался, а пул остался висеть - процесс был жив, но ни одного
        нового домена в чекпойнт больше не попало. Один кривой тег остановил
        прогон на восемьсот доменов.
        """
        try:
            return audit_domain(domain, n_art)
        except Exception as e:                                   # noqa: BLE001
            return {'domain': domain, 'status': 'error', 'error': repr(e)[:200],
                    'sitemap_urls': 0}

    with cf.ThreadPoolExecutor(max_workers=nw) as ex:
        for d in ex.map(safe, doms):
            res.append(d)
            ck.write(json.dumps(d, ensure_ascii=False) + '\n')
            ck.flush()
            os.fsync(ck.fileno())
            print(f"  {d['domain']:28} {d.get('status'):10} "
                  f"статей {d.get('checked', 0):3} | реклама {d.get('pct_ad', 0):3}% | "
                  f"dofollow {d.get('pct_dofollow_of_ad', 0):3}% | не в лист. {d.get('pct_not_in_listings', 0):3}%",
                  flush=True)
    if '--ahrefs' in args:
        res = merge_ahrefs(res, args[args.index('--ahrefs') + 1])
    for d in res:
        d['verdict'] = verdict(d)
    json.dump(res, open('donor-pages-audit.json', 'w'), ensure_ascii=False, indent=1)

    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook(); bold = Font(bold=True)
    w1 = wb.active; w1.title = 'Домены'
    w1.append(['Домен', 'Статус', 'Проверено', '% рекламных', '% dofollow', '% не в листингах',
               'Внеш/статью', 'С пометкой', 'sponsored', 'UR>0 %', f'UR>=5 %', 'Медиана UR',
               'Вердикт', 'Топ внешних хостов'])
    for c in w1[1]: c.font = bold
    for d in sorted(res, key=lambda x: -(x.get('pct_ur_threshold') or x.get('pct_dofollow_of_ad') or 0)):
        w1.append([d['domain'], d.get('status'), d.get('checked'), d.get('pct_ad'),
                   d.get('pct_dofollow_of_ad'), d.get('pct_not_in_listings'),
                   d.get('avg_ext_per_ad'), d.get('ad_labeled'), d.get('sponsored_links'),
                   d.get('pct_ur_positive'), d.get('pct_ur_threshold'), d.get('ur_median'),
                   d.get('verdict'), ', '.join(f'{h}({n})' for h, n in (d.get('top_ext_hosts') or [])[:6])])
    for col, wd in zip('ABCDEFGHIJKLMN', (26, 11, 10, 12, 11, 9, 12, 11, 11, 9, 9, 11, 46, 60)):
        w1.column_dimensions[col].width = wd
    w2 = wb.create_sheet('Статьи')
    w2.append(['Домен', 'URL статьи', 'HTTP', 'Знаков', 'dofollow', 'nofollow', 'sponsored',
               'Внутр. ссылок', 'UR', 'Пометка рекламы', 'Внешние хосты'])
    for c in w2[1]: c.font = bold
    for d in res:
        for a in d.get('articles', []):
            w2.append([d['domain'], a['url'], a.get('http'), a.get('chars'),
                       a.get('ext_dofollow'), a.get('ext_nofollow'), a.get('sponsored'),
                       a.get('internal_links'), a.get('ur'),
                       'да' if a.get('has_ad_label') else '',
                       ', '.join(sorted({l['host'] for l in a.get('ext_links', [])}))[:120]])
    for col, wd in zip('ABCDEFGHIJK', (24, 70, 7, 9, 10, 10, 11, 13, 6, 15, 60)):
        w2.column_dimensions[col].width = wd
    wb.save('donor-pages-audit.xlsx')
    print('\n=> donor-pages-audit.xlsx / .json')
    for d in sorted(res, key=lambda x: -(x.get('pct_dofollow_of_ad') or 0)):
        print(f"  {d['domain']:28} {d.get('verdict')}")


if __name__ == '__main__':
    main()
