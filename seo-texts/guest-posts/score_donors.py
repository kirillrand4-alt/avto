# -*- coding: utf-8 -*-
"""Скоринг доноров Miralinks под B2B-промышленную сеть (компрессоры).

    python3 score_donors.py <выгрузка.xlsx> [-o donors-scored.xlsx]

Формула - консенсус панели 05.08.2026: идеи от 4 ИИ (fable-5, gemini-3.1-pro,
gpt-5.6-terra, grok-4.5) -> 3 судьи-синтезатора (DONOR-CRITERIA-IDEAS.md /
DONOR-CRITERIA-JUDGES.md). Взяты веса судьи fable (самые проработанные пороги),
блок Majestic-когерентности и штрафы за рассогласование метрик - консенсус 3 из 3.

Выход: XLSX с 4 листами (скоринг, корзины, отбракованные, сводка) + печать топа.
"""
import math, os, re, sys

import openpyxl
from openpyxl.styles import Font, PatternFill

# --- тематические группы (мультизначная колонка «Тематика») ---
INDUSTRIAL = ('сельское хозяйство', 'производств', 'строительств', 'транспортн',
              'авто и мото', 'промышленн', 'оборудован', 'энергетик', 'ремонт')
BUSINESS = ('бизнес', 'финанс', 'экономик', 'наука', 'образование')
MEDIA = ('газеты', 'смi', 'сми', 'портал')
# токсичные соседства - отбраковка (консенсус всех линз)
TOXIC = ('казино', 'гэмблинг', 'ставк', 'букмекер', 'займ', 'мфо', 'кредит',
         'эзотер', 'гадан', 'знакомств', 'adult', 'взросл', 'заработок в интернете')
# 309 значений региона: страны, области, города РФ. Не-РФ/СНГ определяем ЧЁРНЫМ списком
# (иначе российские области ловятся как «зарубеж» - баг первого прогона 05.08).
REGION_BAD = ('украина', 'киев', 'одесс', 'харьков', 'львов', 'днепр', 'индонези',
              'сша', 'индия', 'китай', 'турци', 'вьетнам', 'польш', 'герман', 'испани',
              'франц', 'итали', 'великобритан', 'бразили', 'мексик', 'таиланд', 'япони',
              'корея', 'филиппин', 'малайзи', 'египет', 'израил', 'молдов', 'грузи',
              'армени', 'азербайджан', 'узбекистан', 'таджикистан', 'киргиз', 'латви',
              'литв', 'эстони', 'болгари', 'румыни', 'серби', 'чехи', 'венгри')


def num(v, default=None):
    try:
        f = float(v)
        return f if not math.isnan(f) else default
    except (TypeError, ValueError):
        return default


def clamp(x, lo, hi):
    return max(lo, min(hi, x))


def topic_block(topics_raw, region):
    """A: 0-24. Тематика - главный вес: ссылка из промышленной рубрики весит кратно больше."""
    t = str(topics_raw or '').lower()
    parts = [p.strip() for p in t.split(';') if p.strip()]
    if any(k in t for k in INDUSTRIAL):
        a = 24.0
    elif any(k in t for k in BUSINESS):
        a = 15.0
    elif any(k in t for k in MEDIA):
        r = str(region or '').lower()
        a = 10.0 if r and not any(b in r for b in REGION_BAD) and 'снг' not in r else 6.0
    else:
        a = 0.0
    if len(parts) >= 5:          # «сайт обо всём» - тематического ядра нет
        a = min(a, 9.0)
    return a


def traffic_block(T, Ta, P):
    """B: 0-22, логарифмическая шкала (300 -> 0, 30000 -> 22) + штрафы за несогласованность."""
    T = T or 0
    b = 22 * clamp(math.log10(max(T, 1) / 300) / math.log10(100), 0, 1)
    if Ta is None:
        b *= 0.85                # нет данных Ahrefs - лёгкий дисконт
    elif 0 < Ta < 0.1 * T:
        b *= 0.75                # органики почти нет: трафик не поисковый
    if P is not None and P < 200 and T / max(P, 1) > 50:
        b *= 0.5                 # Дзен/Турбо-трафик на мизерном индексе
    return b


def authority_block(iks, dr, da, T):
    """C: 0-16 (ИКС 0-10 + DR/DA 0-6) с обнулением накруток."""
    iks = iks or 0
    c_iks = (0 if iks < 30 else 3 if iks < 60 else 5 if iks < 100 else
             7 if iks < 160 else 8.5 if iks < 300 else 9.5 if iks < 1000 else 10)
    d = dr if dr is not None else da
    if d is None:
        c_dr = 2.0               # пропуск - нейтрально-низко (у трети базы DR нет)
    else:
        c_dr = (0 if d < 10 else 2 if d < 20 else 4 if d < 35 else 5 if d < 50 else 6)
    if d is not None and d >= 45 and (T or 0) < 300:
        c_dr = 0                 # PBN: сильный DR при мёртвом трафике
    if iks >= 300 and (T or 0) < 500:
        c_iks = 3                # дроп с унаследованным ИКС
    return c_iks + c_dr


def index_block(ix, P, S):
    """D: 0-16. Неиндексируемая статья = ссылки не существует."""
    ix = ix if ix is not None else 0
    d = (7 if ix >= 95 else 5 if ix >= 85 else 2 if ix >= 75 else 0)
    P = P or 0
    d += (5 if P >= 300 else 4 if P >= 100 else 2 if P >= 30 else 0)
    r = (S or 0) / max(P, 1)     # доля гостевых в индексе
    d += (4 if r <= 0.05 else 2.5 if r <= 0.15 else 0 if r <= 0.30 else -5)
    return d


def outgoing_block(S, T):
    """E: 0-14. Сайт-конвейер размещений обнуляет вес исходящего."""
    S = S or 0
    e = (14 if S <= 30 else 11 if S <= 100 else 7 if S <= 250 else
         3 if S <= 500 else 0 if S <= 700 else -10)
    if T and S / max(T / 1000, 0.001) > 50:
        e -= 4
    return e


def majestic_block(cf, tf):
    """F: 0-2 + возможный множитель на C. CF/TF - детектор ссылочной схемы."""
    if cf is None or tf is None or tf == 0:
        return 0.75, 1.0
    ratio = cf / tf
    if tf >= 15 and ratio <= 1.8:
        return 2.0, 1.0
    if tf >= 10 and ratio <= 2.2:
        return 1.25, 1.0
    if ratio > 2.5:
        return 0.0, 0.75
    return 0.5, 1.0


def spam_block(sp):
    """H: 0-6. Спам Я - доля (0.16 = 16%), в базе 0..0.20, p90=0.17.
    ИСПРАВЛЕНО 05.08: в первом прогоне колонка была ошибочно сочтена нулевой
    (виновата печать с округлением до целых, а не данные). Ненулевых 9182 из 14338."""
    if sp is None:
        return 1.5           # нет данных - нейтрально-низко
    if sp <= 0.02:
        return 6.0
    if sp <= 0.05:
        return 5.0
    if sp <= 0.10:
        return 3.5
    if sp <= 0.15:
        return 2.0
    if sp <= 0.18:
        return 0.5
    return 0.0               # 18-20%: верхняя граница базы, соседство мусорное


def economy_block(price, excl, region):
    """G: 0-4. Цена - обратный индикатор: дно биржи не проводит редактуру."""
    price = price or 0
    g = (1 if price >= 2500 else 0.75 if price >= 1200 else 0.25 if price >= 500 else 0)
    if str(excl) == '1':
        g += 0.5
    r = str(region or '').lower()
    if r and not any(b in r for b in REGION_BAD) and 'снг' not in r:
        g += 0.5             # конкретный российский регион/город или «Россия»
    return g


def hard_filters(d):
    """Отсечки ДО скоринга -> причина отбраковки или None."""
    lang = str(d.get('Язык') or '').lower()
    region = str(d.get('Регион') or '').lower()
    topics = str(d.get('Тематика') or '').lower()
    T, iks = num(d.get('Трафик'), 0), num(d.get('ИКС'), 0)
    ix, S, P = num(d.get('% индекс Я')), num(d.get('Статей'), 0), num(d.get('Стр. в индексе Я'), 0)
    price = num(d.get('Цена ₽'), 0)
    dr = num(d.get('Ahrefs DR'))
    if lang and lang != 'ru':
        return 'язык не русский'
    if any(r in region for r in REGION_BAD):
        return f'регион вне РФ/СНГ ({d.get("Регион")})'
    if any(k in topics for k in TOXIC):
        return 'токсичная тематика (гэмблинг/займы/эзотерика)'
    if T < 100 and iks < 30:
        return 'мёртвый хост (трафик <100 и ИКС <30)'
    if ix is not None and ix < 70:
        return f'индексация {ix:.0f}% (<70)'
    if S > 700:
        return f'линкопомойка ({S:.0f} размещений)'
    if P and S / max(P, 1) > 0.35:
        return 'гостевые >35% индекса'
    if price < 400:
        return f'цена {price:.0f}₽ (<400, дно биржи)'
    if 0 < P < 25:
        return f'сайт из {P:.0f} страниц'
    if dr is not None and dr >= 45 and T < 300:
        return 'PBN (DR>=45 при трафике <300)'
    if iks >= 500 and T < 500:
        return 'дроп с унаследованным ИКС'
    if len([p for p in topics.split(';') if p.strip()]) >= 7:
        return 'нет тематического ядра (>=7 тематик)'
    sp = num(d.get('Спам Я'))
    if sp is not None and sp >= 0.18:
        return f'спам Яндекса {sp*100:.0f}% (>=18%, верх базы)'
    return None


def score_row(d):
    T, Ta = num(d.get('Трафик'), 0), num(d.get('Ahrefs трафик'))
    P, S = num(d.get('Стр. в индексе Я')), num(d.get('Статей'), 0)
    A = topic_block(d.get('Тематика'), d.get('Регион'))
    B = traffic_block(T, Ta, P)
    C = authority_block(num(d.get('ИКС'), 0), num(d.get('Ahrefs DR')), num(d.get('Moz DA')), T)
    D = index_block(num(d.get('% индекс Я')), P, S)
    E = outgoing_block(S, T)
    F, c_mult = majestic_block(num(d.get('CF')), num(d.get('TF')))
    C *= c_mult
    G = economy_block(num(d.get('Цена ₽'), 0), d.get('Эксклюзив'), d.get('Регион'))
    H = spam_block(num(d.get('Спам Я')))
    total = A + B + C + D + E + F + G + H
    return dict(A=A, B=B, C=C, D=D, E=E, F=F, G=G, H=H, SCORE=total)


def basket(d, s):
    """Корзины сегментации (не только топ формулы) - консенсус судей."""
    topics = str(d.get('Тематика') or '').lower()
    T, iks = num(d.get('Трафик'), 0), num(d.get('ИКС'), 0)
    S = num(d.get('Статей'), 0)
    ind = any(k in topics for k in INDUSTRIAL)
    if ind and T >= 1000 and S <= 250:
        return '1-отраслевое ядро'
    reg = str(d.get('Регион') or '').lower()
    if any(k in topics for k in MEDIA) and reg and not any(b in reg for b in REGION_BAD) \
            and 'снг' not in reg and T >= 5000 and iks >= 100:
        return '2-региональные СМИ'
    if T >= 30000 and iks >= 300:
        return '3-трафиковые тяжеловесы'
    if ('авто и мото' in topics or 'сельское хозяйство' in topics) and s >= 50:
        return '4-смежные потребители'
    if 52 <= s < 62:
        return '5-эксперимент'
    return '-'


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else 'donors.xlsx'
    out = sys.argv[sys.argv.index('-o') + 1] if '-o' in sys.argv else 'donors-scored.xlsx'
    ws = openpyxl.load_workbook(src, read_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    data = [dict(zip(hdr, r)) for r in rows[1:] if r[0]]
    print(f'загружено доноров: {len(data)}')

    passed, rejected = [], []
    for d in data:
        why = hard_filters(d)
        if why:
            rejected.append((d, why))
            continue
        sc = score_row(d)
        d['_score'] = sc
        d['_basket'] = basket(d, sc['SCORE'])
        passed.append(d)
    passed.sort(key=lambda x: -x['_score']['SCORE'])
    print(f'прошли отсечки: {len(passed)} | отбраковано: {len(rejected)}')

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    green = PatternFill('solid', fgColor='C6EFCE')
    yellow = PatternFill('solid', fgColor='FFEB9C')

    w1 = wb.active; w1.title = 'Скоринг'
    cols = ['#', 'Домен', 'Название', 'SCORE', 'Корзина', 'Тематика(A)', 'Трафик(B)',
            'Траст(C)', 'Индекс(D)', 'Исходящие(E)', 'Majestic(F)', 'Эконом(G)', 'Спам(H)',
            'Трафик', 'Ahrefs траф', 'ИКС', 'DR', '% индекс', 'Стр. индекс',
            'Статей', 'Спам Я %', 'Цена ₽', 'Регион', 'Тематика']
    w1.append(cols)
    for c in w1[1]: c.font = bold
    for i, d in enumerate(passed, 1):
        s = d['_score']
        w1.append([i, d['Домен'], d['Название'], round(s['SCORE'], 1), d['_basket'],
                   round(s['A'], 1), round(s['B'], 1), round(s['C'], 1), round(s['D'], 1),
                   round(s['E'], 1), round(s['F'], 1), round(s['G'], 1), round(s['H'], 1),
                   num(d['Трафик'], 0), num(d.get('Ahrefs трафик')), num(d['ИКС'], 0),
                   num(d.get('Ahrefs DR')), num(d.get('% индекс Я')),
                   num(d.get('Стр. в индексе Я')), num(d.get('Статей'), 0),
                   round((num(d.get('Спам Я')) or 0) * 100, 1),
                   num(d.get('Цена ₽'), 0), d.get('Регион'), str(d.get('Тематика'))[:80]])
        fill = green if s['SCORE'] >= 72 else (yellow if s['SCORE'] >= 62 else None)
        if fill:
            for c in w1[w1.max_row]: c.fill = fill
    for col, wd in zip('ABCDEFGHIJKLMNOPQRSTUVWX',
                       (5, 26, 30, 8, 20, 10, 9, 8, 9, 11, 10, 9, 8, 10, 11, 8, 6, 9, 11, 8, 9, 8, 12, 50)):
        w1.column_dimensions[col].width = wd
    w1.freeze_panes = 'A2'

    w2 = wb.create_sheet('Корзины')
    w2.append(['Корзина', 'Доноров', 'Медиана SCORE', 'Медиана цены', 'Примеры (топ-8)'])
    for c in w2[1]: c.font = bold
    import statistics
    bs = {}
    for d in passed:
        bs.setdefault(d['_basket'], []).append(d)
    for b in sorted(bs):
        items = sorted(bs[b], key=lambda x: -x['_score']['SCORE'])
        w2.append([b, len(items),
                   round(statistics.median([x['_score']['SCORE'] for x in items]), 1),
                   round(statistics.median([num(x.get('Цена ₽'), 0) for x in items])),
                   ', '.join(x['Домен'] for x in items[:8])])
    for col, wd in zip('ABCDE', (24, 10, 14, 14, 90)):
        w2.column_dimensions[col].width = wd

    w3 = wb.create_sheet('Отбраковка')
    w3.append(['Домен', 'Причина', 'Трафик', 'ИКС', 'Цена'])
    for c in w3[1]: c.font = bold
    reasons = {}
    for d, why in rejected:
        key = re.sub(r'\d+', 'N', why)
        reasons[key] = reasons.get(key, 0) + 1
    for d, why in rejected[:3000]:
        w3.append([d['Домен'], why, num(d.get('Трафик'), 0), num(d.get('ИКС'), 0), num(d.get('Цена ₽'), 0)])
    for col, wd in zip('ABCDE', (28, 40, 10, 8, 8)):
        w3.column_dimensions[col].width = wd

    w4 = wb.create_sheet('Сводка')
    w4.append(['Показатель', 'Значение'])
    for c in w4[1]: c.font = bold
    green_n = sum(1 for d in passed if d['_score']['SCORE'] >= 72)
    yellow_n = sum(1 for d in passed if 62 <= d['_score']['SCORE'] < 72)
    exp_n = sum(1 for d in passed if 52 <= d['_score']['SCORE'] < 62)
    for k, v in [('Всего в выгрузке', len(data)), ('Прошли hard-фильтры', len(passed)),
                 ('Отбраковано', len(rejected)),
                 ('Корзина A (SCORE>=72, покупаем)', green_n),
                 ('Корзина B (62-71, после проверок)', yellow_n),
                 ('Эксперимент (52-61)', exp_n)]:
        w4.append([k, v])
    w4.append(['', ''])
    w4.append(['Причины отбраковки', 'шт'])
    for r, n in sorted(reasons.items(), key=lambda kv: -kv[1]):
        w4.append([r, n])
    w4.column_dimensions['A'].width = 46; w4.column_dimensions['B'].width = 12
    wb.save(out)

    print(f'\nКорзина A (>=72): {green_n} | B (62-71): {yellow_n} | эксперимент (52-61): {exp_n}')
    print('\nТОП-25:')
    for i, d in enumerate(passed[:25], 1):
        s = d['_score']
        print(f"{i:3}. {s['SCORE']:5.1f} {d['Домен'][:30]:30} траф {num(d['Трафик'],0):>7.0f} "
              f"ИКС {num(d['ИКС'],0):>6.0f} ст.{num(d.get('Статей'),0):>5.0f} "
              f"{str(d.get('Цена ₽')):>6}₽ | {d['_basket'][:20]:20} | {str(d.get('Тематика'))[:45]}")
    print('\nПричины отбраковки (топ-10):')
    for r, n in sorted(reasons.items(), key=lambda kv: -kv[1])[:10]:
        print(f'  {n:6} {r}')
    print(f'\n=> {out}')


if __name__ == '__main__':
    main()
