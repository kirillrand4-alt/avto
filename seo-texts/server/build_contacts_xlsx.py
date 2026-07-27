# -*- coding: utf-8 -*-
"""Сборка XLSX с НОВЫМИ контактами по двум базам (запускается НА СЕРВЕРЕ).

Требования владельца:
  - строка на контакт;
  - телефоны закупщиков идут первыми и подсвечены;
  - колонка «Контакт» — КЛИКАБЕЛЬНАЯ ссылка на страницу-источник;
  - колонки: ИНН, компания, тип, контакт, ФИО+роль, источник, сайт, регион, ОКВЭД, выручка;
  - в выгрузку попадают только контакты, которых НЕТ в исходной базе продажников
    (дедуп по нормализованным 10 цифрам телефона и по email).

Источник данных — enrich.db (таблицы companies / phone_contacts / emails), то есть
результат ШТАТНОГО конвейера, а не самописных парсеров.

Запуск:  python build_contacts_xlsx.py [выходной_каталог]
"""
import json
import os
import re
import sqlite3
import sys

DB = r'C:\sender\enrich.db'
SALES = r'C:\sender\_ops\sales_base.json'
CORE = r'C:\seostat\drop\drop-storage\centrifugal-core-inns.txt'
OUT_DIR_DEFAULT = r'C:\seostat\drop\drop-storage'

INN_RE = re.compile(r'\b(\d{10}|\d{12})\b')
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')

# роль/источник, означающие закупщика — такие строки первыми и подсвечены
PROC_RE = re.compile(r'закуп|снабж|тендер|постав', re.I)


# добавочный: конвейер СОХРАНЯЕТ их по промпту extract_roles
# («+7 342 292-14-60 доб.122 (отдел продаж)»), поэтому режем хвост ДО нормализации,
# иначе валидный номер выглядит как 14 цифр и вылетает в мусор.
EXT_RE = re.compile(r'(доб|добав|вн\.?|ext|extension|#|,)', re.I)


def norm_phone(raw):
    """Телефон -> 10 цифр или None. Отсекает ОГРН/счета/заглушки (грабля владельца)."""
    if not raw:
        return None
    s = str(raw)
    # отрезаем добавочный и всё, что после него (в т.ч. «(отдел продаж)»)
    m = EXT_RE.search(s)
    if m:
        s = s[:m.start()]
    d = re.sub(r'\D', '', s)
    if len(d) == 11 and d[0] in '78':
        d = d[1:]
    if len(d) != 10:
        # ОГРН (13), счёт (20), обрезки — НЕ телефон
        return None
    if d[0] in '012':
        return None                 # не бывает российских кодов на 0/1/2 (ловит хвосты ОГРН)
    if d[:3] == '000':
        return None
    if len(set(d)) == 1:            # 0000000000 / 9999999999
        return None
    if d[3:] == '0000000':          # +7 (XXX) 000-00-00
        return None
    return d


def fmt_phone(d10):
    return f'+7 ({d10[:3]}) {d10[3:6]}-{d10[6:8]}-{d10[8:]}'


def norm_email(raw):
    if not raw:
        return None
    e = str(raw).strip().lower()
    return e if EMAIL_RE.fullmatch(e) else None


def load_sales_known():
    """Что УЖЕ есть у продажников: множества телефонов(10 цифр) и email."""
    phones, emails, inns = set(), set(), []
    if not os.path.exists(SALES):
        return phones, emails, inns
    d = json.load(open(SALES, encoding='utf-8'))
    items = []
    if isinstance(d, dict):
        for v in d.values():
            items.extend(v or [])
    else:
        items = d
    seen = set()
    for r in items:
        if not isinstance(r, dict):
            continue
        i = str(r.get('inn') or '').strip()
        m = INN_RE.search(i) if i else None
        if m and m.group(1) not in seen:
            seen.add(m.group(1))
            inns.append(m.group(1))
        for p in (r.get('phones') or []):
            n = norm_phone(p)
            if n:
                phones.add(n)
        for e in (r.get('emails') or []):
            n = norm_email(e)
            if n:
                emails.add(n)
    return phones, emails, inns


def load_core_inns():
    out = []
    if not os.path.exists(CORE):
        return out
    for ln in open(CORE, encoding='utf-8', errors='replace'):
        m = INN_RE.search(ln)
        if m:
            out.append(m.group(1))
    return sorted(set(out))


def fetch(cx, inns):
    """Компании + контакты по набору ИНН."""
    cx.execute('drop table if exists temp.t_inn')
    cx.execute('create temp table t_inn(inn text primary key)')
    cx.executemany('insert or ignore into t_inn values(?)', [(i,) for i in inns])

    comp = {}
    for r in cx.execute(
            'select c.inn, c.name, c.site, c.region, c.okved, c.revenue_rub, '
            'c.activity, c.is_competitor from companies c join t_inn t on t.inn=c.inn'):
        comp[r[0]] = {'name': r[1] or '', 'site': r[2] or '', 'region': r[3] or '',
                      'okved': r[4] or '', 'revenue': r[5] or '',
                      'activity': r[6] or '', 'competitor': r[7] or 0}

    phones = list(cx.execute(
        'select p.inn, p.phone, p.person, p.role, p.source, p.source_url '
        'from phone_contacts p join t_inn t on t.inn=p.inn'))
    emails = list(cx.execute(
        'select e.inn, e.email, e.person, e.role, e.source, e.source_url '
        'from emails e join t_inn t on t.inn=e.inn'))
    return comp, phones, emails


def build_rows(comp, phones, emails, known_phones, known_emails):
    """Строки выгрузки: только НОВЫЕ контакты, закупщики первыми."""
    rows, seen = [], set()
    stat = {'телефонов_всего': len(phones), 'email_всего': len(emails),
            'мусорных_телефонов': 0, 'уже_у_продажников': 0, 'дублей': 0}

    for inn, phone, person, role, source, url in phones:
        n = norm_phone(phone)
        if not n:
            stat['мусорных_телефонов'] += 1
            continue
        if n in known_phones:
            stat['уже_у_продажников'] += 1
            continue
        k = (inn, 'тел', n)
        if k in seen:
            stat['дублей'] += 1
            continue
        seen.add(k)
        c = comp.get(inn, {})
        rows.append({
            'inn': inn, 'company': c.get('name', ''), 'kind': 'телефон',
            'contact': fmt_phone(n), 'who': ' — '.join(x for x in (person, role) if x),
            'source': source or '', 'url': url or '', 'site': c.get('site', ''),
            'region': c.get('region', ''), 'okved': c.get('okved', ''),
            'revenue': c.get('revenue', ''),
            'proc': bool(PROC_RE.search(f'{role or ""} {source or ""}')),
        })

    for inn, email, person, role, source, url in emails:
        n = norm_email(email)
        if not n:
            continue
        if n in known_emails:
            stat['уже_у_продажников'] += 1
            continue
        k = (inn, 'email', n)
        if k in seen:
            stat['дублей'] += 1
            continue
        seen.add(k)
        c = comp.get(inn, {})
        rows.append({
            'inn': inn, 'company': c.get('name', ''), 'kind': 'email',
            'contact': n, 'who': ' — '.join(x for x in (person, role) if x),
            'source': source or '', 'url': url or '', 'site': c.get('site', ''),
            'region': c.get('region', ''), 'okved': c.get('okved', ''),
            'revenue': c.get('revenue', ''),
            'proc': bool(PROC_RE.search(f'{role or ""} {source or ""}')),
        })

    # закупщики первыми; внутри групп — по компании, телефоны раньше email
    rows.sort(key=lambda r: (not r['proc'], r['company'],
                             0 if r['kind'] == 'телефон' else 1, r['contact']))
    stat['строк_в_выгрузке'] = len(rows)
    stat['из_них_закупщики'] = sum(1 for r in rows if r['proc'])
    stat['с_кликабельной_ссылкой'] = sum(1 for r in rows if r['url'])
    return rows, stat


HEAD = ['ИНН', 'Компания', 'Тип', 'Контакт', 'ФИО и роль', 'Источник',
        'Сайт', 'Регион', 'ОКВЭД', 'Выручка, ₽']


def write_xlsx(path, rows, title):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    head_fill = PatternFill('solid', fgColor='1F3864')
    head_font = Font(color='FFFFFF', bold=True)
    proc_fill = PatternFill('solid', fgColor='FFF2CC')     # подсветка закупщиков
    link_font = Font(color='0563C1', underline='single')

    ws.append(HEAD)
    for c in ws[1]:
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(vertical='center')

    for r in rows:
        ws.append([r['inn'], r['company'], r['kind'], r['contact'], r['who'],
                   r['source'], r['site'], r['region'], r['okved'], r['revenue']])
        i = ws.max_row
        # «Контакт» — кликабельная ссылка на страницу-источник
        if r['url']:
            cell = ws.cell(row=i, column=4)
            cell.hyperlink = r['url']
            cell.font = link_font
        if r['site']:
            s = r['site'] if str(r['site']).startswith('http') else 'http://' + str(r['site'])
            cell = ws.cell(row=i, column=7)
            cell.hyperlink = s
            cell.font = link_font
        if r['proc']:
            for col in range(1, len(HEAD) + 1):
                ws.cell(row=i, column=col).fill = proc_fill

    widths = [14, 46, 9, 24, 34, 20, 30, 20, 12, 18]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEAD))}{ws.max_row}'
    wb.save(path)


def main():
    out_dir = sys.argv[1] if len(sys.argv) > 1 else OUT_DIR_DEFAULT
    os.makedirs(out_dir, exist_ok=True)
    known_phones, known_emails, sales_inns = load_sales_known()
    core_inns = load_core_inns()

    cx = sqlite3.connect(f'file:{DB}?mode=ro', uri=True, timeout=60)
    report = {'известно_у_продажников': {'телефонов': len(known_phones),
                                         'email': len(known_emails)}}

    for label, inns, fname in (
            ('продажники', sales_inns, 'sales-new-contacts.xlsx'),
            ('ядро центробежных', core_inns, 'core-new-contacts.xlsx')):
        comp, ph, em = fetch(cx, inns)
        rows, stat = build_rows(comp, ph, em, known_phones, known_emails)
        path = os.path.join(out_dir, fname)
        write_xlsx(path, rows, label)
        stat['ИНН_в_списке'] = len(inns)
        stat['компаний_с_контактами'] = len({r['inn'] for r in rows})
        stat['файл'] = path
        report[label] = stat

    cx.close()
    print(json.dumps(report, ensure_ascii=False))


if __name__ == '__main__':
    main()
